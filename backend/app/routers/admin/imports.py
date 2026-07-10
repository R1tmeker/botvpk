from __future__ import annotations

import csv
import io
import json
import secrets
from datetime import timedelta
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from redis.asyncio import Redis
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ...config import Settings, get_settings
from ...database import get_db_session
from ...dependencies.auth import CurrentUser, require_role
from ...models import Squad, User
from ...roles import ALL_INTERNAL_ROLES, RoleLevel
from ...schemas.product import ImportChange, ImportCommitResult, ImportPreview, ImportRowIssue
from ...services.sessions import delete_user_sessions
from ...utils.audit import record_audit, utcnow

router = APIRouter(prefix="/admin/imports", tags=["admin:imports"])
IMPORT_FIELDS = ("telegram_id", "username", "full_name", "squad_id", "role_code", "status_code")
HEADER_ALIASES = {
    "telegram_id": "telegram_id",
    "telegram": "telegram_id",
    "телеграм": "telegram_id",
    "tg_id": "telegram_id",
    "username": "username",
    "логин": "username",
    "full_name": "full_name",
    "фио": "full_name",
    "имя": "full_name",
    "squad_id": "squad_id",
    "отделение": "squad_id",
    "role_code": "role_code",
    "роль": "role_code",
    "status_code": "status_code",
    "статус": "status_code",
}


def require_profile(current_user: CurrentUser) -> int:
    if current_user.user_id is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="User profile is required.")
    return current_user.user_id


def normalize_header(value: Any) -> str:
    return HEADER_ALIASES.get(str(value or "").strip().casefold(), str(value or "").strip().casefold())


def read_import_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = filename.casefold().rsplit(".", 1)[-1]
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        return [
            {normalize_header(key): value for key, value in row.items() if key is not None}
            for row in reader
        ]
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [normalize_header(value) for value in next(rows, ())]
        result: list[dict[str, Any]] = []
        for values in rows:
            if any(isinstance(value, str) and value.startswith("=") for value in values):
                raise ValueError("Formulas are not allowed in import files.")
            if not any(value not in (None, "") for value in values):
                continue
            result.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
        workbook.close()
        return result
    raise ValueError("Only CSV and XLSX files are supported.")


def normalize_integer(value: Any, field: str) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except ValueError as exc:
        raise ValueError(f"{field} must be an integer") from exc


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    telegram_id = normalize_integer(row.get("telegram_id"), "telegram_id")
    if telegram_id is None or telegram_id <= 0:
        raise ValueError("telegram_id is required and must be positive")
    full_name = str(row.get("full_name") or "").strip()
    if not full_name:
        raise ValueError("full_name is required")
    role_code = str(row.get("role_code") or "PARTICIPANT").strip().upper()
    if role_code not in ALL_INTERNAL_ROLES:
        raise ValueError(f"Unsupported role_code: {role_code}")
    status_code = str(row.get("status_code") or "ACTIVE").strip().upper()
    if status_code not in {"ACTIVE", "ARCHIVED"}:
        raise ValueError(f"Unsupported status_code: {status_code}")
    username = str(row.get("username") or "").strip().lstrip("@") or None
    return {
        "telegram_id": telegram_id,
        "username": username,
        "full_name": full_name,
        "squad_id": normalize_integer(row.get("squad_id"), "squad_id"),
        "role_code": role_code,
        "status_code": status_code,
    }


def user_snapshot(user: User) -> dict[str, Any]:
    return {field: getattr(user, field) for field in IMPORT_FIELDS}


def redis_client(settings: Settings) -> Redis:
    if not settings.redis_url:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Redis is required for imports.")
    return Redis.from_url(settings.redis_url, decode_responses=True)


@router.post("/preview", response_model=ImportPreview)
async def preview_import(
    upload: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_role(RoleLevel.ADMIN)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> ImportPreview:
    owner_id = require_profile(current_user)
    content = await upload.read(5 * 1024 * 1024 + 1)
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Import file exceeds 5 MB.")
    try:
        raw_rows = read_import_rows(upload.filename or "", content)
    except (UnicodeDecodeError, csv.Error, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)) from exc
    if not raw_rows or len(raw_rows) > 5000:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Import must contain between 1 and 5000 data rows.",
        )

    errors: list[ImportRowIssue] = []
    normalized: list[tuple[int, dict[str, Any]]] = []
    seen_telegram_ids: set[int] = set()
    for row_number, row in enumerate(raw_rows, start=2):
        try:
            item = normalize_import_row(row)
            if item["telegram_id"] in seen_telegram_ids:
                raise ValueError("Duplicate telegram_id in import file")
            if item["role_code"] == "SUPER_ADMIN" and current_user.role_level < RoleLevel.SUPER_ADMIN:
                raise ValueError("Only SUPER_ADMIN can import another SUPER_ADMIN")
            seen_telegram_ids.add(item["telegram_id"])
            normalized.append((row_number, item))
        except ValueError as exc:
            errors.append(ImportRowIssue(row=row_number, message=str(exc)))

    squad_ids = {item["squad_id"] for _, item in normalized if item["squad_id"] is not None}
    existing_squad_ids = set((await session.scalars(select(Squad.id).where(Squad.id.in_(squad_ids)))).all()) if squad_ids else set()
    valid_rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, item in normalized:
        if item["squad_id"] is not None and item["squad_id"] not in existing_squad_ids:
            errors.append(ImportRowIssue(row=row_number, message=f"Unknown squad_id: {item['squad_id']}"))
        else:
            valid_rows.append((row_number, item))

    existing_users = {
        user.telegram_id: user
        for user in (
            await session.scalars(select(User).where(User.telegram_id.in_([item["telegram_id"] for _, item in valid_rows])))
        ).all()
    } if valid_rows else {}
    changes: list[ImportChange] = []
    stored_changes: list[dict[str, Any]] = []
    for row_number, after in valid_rows:
        existing = existing_users.get(after["telegram_id"])
        before = user_snapshot(existing) if existing else None
        action = "CREATE" if existing is None else ("UNCHANGED" if before == after else "UPDATE")
        change = ImportChange(
            row=row_number,
            action=action,
            identity=str(after["telegram_id"]),
            before=before,
            after=after,
        )
        changes.append(change)
        stored_changes.append(change.model_dump(mode="json"))

    preview_id = secrets.token_urlsafe(18)
    expires_at = utcnow() + timedelta(hours=1)
    payload = {
        "owner_id": owner_id,
        "has_errors": bool(errors),
        "changes": stored_changes,
    }
    redis = redis_client(settings)
    try:
        await redis.setex(f"botvpk:import:{preview_id}", 3600, json.dumps(payload, ensure_ascii=False))
    finally:
        await redis.aclose()
    return ImportPreview(
        preview_id=preview_id,
        total_rows=len(raw_rows),
        create_count=sum(item.action == "CREATE" for item in changes),
        update_count=sum(item.action == "UPDATE" for item in changes),
        unchanged_count=sum(item.action == "UNCHANGED" for item in changes),
        errors=errors,
        changes=changes,
        expires_at=expires_at,
    )


@router.post("/{preview_id}/commit", response_model=ImportCommitResult)
async def commit_import(
    preview_id: str,
    current_user: CurrentUser = Depends(require_role(RoleLevel.ADMIN)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> ImportCommitResult:
    owner_id = require_profile(current_user)
    redis = redis_client(settings)
    try:
        raw = await redis.get(f"botvpk:import:{preview_id}")
        if raw is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import preview expired or not found.")
        preview = json.loads(raw)
        if preview["owner_id"] != owner_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Import preview belongs to another user.")
        if preview["has_errors"]:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Import preview contains errors.")

        audit_before: list[dict[str, Any]] = []
        audit_after: list[dict[str, Any]] = []
        affected_user_ids: list[int] = []
        created = updated = 0
        for change in preview["changes"]:
            if change["action"] == "UNCHANGED":
                continue
            after = change["after"]
            user = await session.scalar(
                select(User).where(User.telegram_id == after["telegram_id"]).with_for_update()
            )
            if change["action"] == "CREATE":
                if user is not None:
                    raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="User data changed after preview.")
                user = User(**after)
                session.add(user)
                await session.flush()
                before = None
                created += 1
            else:
                if user is None or user_snapshot(user) != change["before"]:
                    raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="User data changed after preview.")
                before = user_snapshot(user)
                for field, value in after.items():
                    setattr(user, field, value)
                user.token_version += 1
                user.updated_at = utcnow()
                updated += 1
            affected_user_ids.append(user.id)
            audit_before.append({"id": user.id, "before": before})
            audit_after.append({"id": user.id, "after": after, "operation": change["action"]})

        audit = await record_audit(
            session,
            user_id=owner_id,
            action_code="admin.import.commit",
            entity_name="users",
            old_value=audit_before,
            new_value=audit_after,
            comment=f"preview_id={preview_id}",
        )
        await session.flush()
        audit_batch_id = audit.id
        await session.commit()
        await redis.delete(f"botvpk:import:{preview_id}")
    finally:
        await redis.aclose()
    for user_id in affected_user_ids:
        await delete_user_sessions(settings, user_id)
    return ImportCommitResult(created=created, updated=updated, audit_batch_id=audit_batch_id)
