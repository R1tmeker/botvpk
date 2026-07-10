from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ...config import Settings, get_settings
from ...database import get_db_session
from ...dependencies.auth import CurrentUser, require_role
from ...models import NotificationPreference, Squad, User
from ...roles import RoleLevel, ROLE_LEVELS
from ...schemas.core import UserRead, UserUpdate
from ...schemas.product import AdminUsersBulkResult, AdminUsersBulkUpdate
from ...services.auth_security import bump_token_version
from ...services.sessions import delete_user_sessions
from ...utils.audit import model_snapshot, record_audit

router = APIRouter(prefix="/admin/users", tags=["admin:users"])

ROSTER_ROLE_CODES = (
    "PARTICIPANT",
    "DEPUTY_SQUAD_COMMANDER",
    "SQUAD_COMMANDER",
    "DEPUTY_PLATOON_COMMANDER",
    "PLATOON_COMMANDER",
    "ADMIN",
    "SUPER_ADMIN",
    "USER_PENDING",
    "CANDIDATE",
)

ROLE_LABELS_RU = {
    "PUBLIC_USER": "Новый пользователь",
    "CANDIDATE": "Кандидат",
    "USER_PENDING": "Ожидает привязки",
    "PARTICIPANT": "Участник",
    "DEPUTY_SQUAD_COMMANDER": "Зам. командира отделения",
    "SQUAD_COMMANDER": "Командир отделения",
    "DEPUTY_PLATOON_COMMANDER": "Зам. командира взвода",
    "PLATOON_COMMANDER": "Командир взвода",
    "ADMIN": "Администратор",
    "SUPER_ADMIN": "Супер-администратор",
}

EXPORT_COLUMNS = ["ФИО", "Отделение", "Роль", "Статус", "Telegram", "Телефон", "Дата рождения", "Дата привязки"]

LEAD_ROLE_TO_SQUAD_FIELD = {
    "SQUAD_COMMANDER": "commander_user_id",
    "DEPUTY_SQUAD_COMMANDER": "deputy_user_id",
}
SQUAD_FIELD_TO_LEAD_ROLE = {value: key for key, value in LEAD_ROLE_TO_SQUAD_FIELD.items()}
NOTIFICATION_CATEGORIES = ("SCHEDULE", "ATTENDANCE", "NORMATIVES", "ANNOUNCEMENTS", "APPEALS", "SYSTEM")


async def clear_user_squad_lead_refs(
    session: AsyncSession,
    user: User,
    *,
    keep_squad_id: int | None = None,
    keep_field: str | None = None,
) -> None:
    rows = list(
        (
            await session.scalars(
                select(Squad).where(
                    (Squad.commander_user_id == user.id) | (Squad.deputy_user_id == user.id)
                )
            )
        ).all()
    )
    for squad in rows:
        if squad.commander_user_id == user.id and not (squad.id == keep_squad_id and keep_field == "commander_user_id"):
            squad.commander_user_id = None
        if squad.deputy_user_id == user.id and not (squad.id == keep_squad_id and keep_field == "deputy_user_id"):
            squad.deputy_user_id = None


async def sync_user_squad_leadership(session: AsyncSession, user: User) -> None:
    field_name = LEAD_ROLE_TO_SQUAD_FIELD.get(user.role_code)
    if field_name is None:
        await clear_user_squad_lead_refs(session, user)
        return
    if user.squad_id is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Select a squad before assigning squad commander roles.",
        )
    squad = await session.get(Squad, user.squad_id)
    if squad is None or not squad.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Squad not found or inactive.")
    if field_name == "commander_user_id" and squad.deputy_user_id == user.id:
        squad.deputy_user_id = None
    if field_name == "deputy_user_id" and squad.commander_user_id == user.id:
        squad.commander_user_id = None
    await clear_user_squad_lead_refs(session, user, keep_squad_id=squad.id, keep_field=field_name)
    previous_user_id = getattr(squad, field_name)
    if previous_user_id is not None and previous_user_id != user.id:
        previous = await session.get(User, previous_user_id)
        if previous is not None and previous.role_code == SQUAD_FIELD_TO_LEAD_ROLE[field_name]:
            previous.role_code = "PARTICIPANT"
    setattr(squad, field_name, user.id)


def _build_users_query(
    current_user: CurrentUser,
    squad_id: int | None,
    role_code: str | None,
    status_code: str | None,
    search: str | None,
    exclude_public: bool,
):
    statement = select(User).order_by(User.squad_id.nullslast(), User.full_name)
    if current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER:
        if current_user.squad_id is None:
            return None
        statement = statement.where(User.squad_id == current_user.squad_id)
    elif squad_id is not None:
        statement = statement.where(User.squad_id == squad_id)
    if exclude_public:
        statement = statement.where(User.role_code != "PUBLIC_USER")
    if role_code is not None:
        statement = statement.where(User.role_code == role_code)
    if status_code is not None:
        statement = statement.where(User.status_code == status_code)
    else:
        statement = statement.where(User.status_code != "ARCHIVED")
    if search:
        escaped = search.strip().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        term = f"%{escaped}%"
        statement = statement.where(User.full_name.ilike(term) | User.username.ilike(term))
    return statement


@router.get("", response_model=list[UserRead])
async def admin_users(
    squad_id: int | None = None,
    role_code: str | None = None,
    status_code: str | None = None,
    search: str | None = None,
    exclude_public: bool = True,
    limit: int = 200,
    offset: int = 0,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> list[User]:
    statement = _build_users_query(current_user, squad_id, role_code, status_code, search, exclude_public)
    if statement is None:
        return []
    statement = statement.offset(max(0, offset)).limit(min(max(1, limit), 500))
    return list((await session.scalars(statement)).all())


def _user_to_row(user: User, squad_map: dict[int, str]) -> list[str]:
    squad = squad_map.get(user.squad_id or -1, "—") if user.squad_id else "—"
    return [
        user.full_name,
        squad,
        ROLE_LABELS_RU.get(user.role_code, user.role_code),
        user.status_code,
        f"@{user.username}" if user.username else "—",
        user.phone or "—",
        user.birth_date.strftime("%d.%m.%Y") if user.birth_date else "—",
        user.linked_at.strftime("%d.%m.%Y") if user.linked_at else "—",
    ]


@router.get("/export.csv")
async def export_users_csv(
    squad_id: int | None = None,
    role_code: str | None = None,
    status_code: str | None = None,
    search: str | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    statement = _build_users_query(current_user, squad_id, role_code, status_code, search, exclude_public=True)
    users = [] if statement is None else list((await session.scalars(statement)).all())
    squad_map = {s.id: s.name for s in (await session.scalars(select(Squad))).all()}
    await record_audit(session, user_id=current_user.user_id, action_code="users.export", entity_name="users", new_value={"format": "csv", "count": len(users)})
    await session.commit()
    output = io.StringIO()
    output.write("﻿")  # BOM для кириллицы в Excel
    writer = csv.writer(output, delimiter=";")
    writer.writerow(EXPORT_COLUMNS)
    for user in users:
        writer.writerow(_user_to_row(user, squad_map))
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="roster.csv"'},
    )


@router.get("/export.xlsx")
async def export_users_xlsx(
    squad_id: int | None = None,
    role_code: str | None = None,
    status_code: str | None = None,
    search: str | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="openpyxl not installed.") from exc
    statement = _build_users_query(current_user, squad_id, role_code, status_code, search, exclude_public=True)
    users = [] if statement is None else list((await session.scalars(statement)).all())
    squad_map = {s.id: s.name for s in (await session.scalars(select(Squad))).all()}
    await record_audit(session, user_id=current_user.user_id, action_code="users.export", entity_name="users", new_value={"format": "xlsx", "count": len(users)})
    await session.commit()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Состав"
    header_fill = PatternFill(fill_type="solid", fgColor="1a2f5a")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, user in enumerate(users, 2):
        for col_idx, value in enumerate(_user_to_row(user, squad_map), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="roster.xlsx"'},
    )


@router.post("/export.csv/send", status_code=200)
async def export_users_csv_send(
    squad_id: int | None = None,
    search: str | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Generate CSV roster and send it to the requester via Telegram bot DM."""
    from ...background import _get_bot
    from aiogram.types import BufferedInputFile

    statement = _build_users_query(current_user, squad_id, None, None, search, exclude_public=True)
    users = [] if statement is None else list((await session.scalars(statement)).all())
    squad_map = {s.id: s.name for s in (await session.scalars(select(Squad))).all()}
    output = io.StringIO()
    output.write("﻿")  # BOM for Cyrillic
    writer = csv.writer(output, delimiter=";")
    writer.writerow(EXPORT_COLUMNS)
    for user in users:
        writer.writerow(_user_to_row(user, squad_map))
    data = output.getvalue().encode("utf-8")
    bot = _get_bot(settings)
    await bot.send_document(
        current_user.telegram_id,
        BufferedInputFile(data, filename="roster.csv"),
        caption=f"Состав ВПК Звезда — {len(users)} чел.",
    )
    return {"sent": True, "count": len(users)}


@router.post("/export.xlsx/send", status_code=200)
async def export_users_xlsx_send(
    squad_id: int | None = None,
    search: str | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Generate XLSX roster and send it to the requester via Telegram bot DM."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="openpyxl not installed.") from exc
    from ...background import _get_bot
    from aiogram.types import BufferedInputFile

    statement = _build_users_query(current_user, squad_id, None, None, search, exclude_public=True)
    users = [] if statement is None else list((await session.scalars(statement)).all())
    squad_map = {s.id: s.name for s in (await session.scalars(select(Squad))).all()}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Состав"
    header_fill = PatternFill(fill_type="solid", fgColor="1a2f5a")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, user in enumerate(users, 2):
        for col_idx, value in enumerate(_user_to_row(user, squad_map), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    bot = _get_bot(settings)
    await bot.send_document(
        current_user.telegram_id,
        BufferedInputFile(data, filename="roster.xlsx"),
        caption=f"Состав ВПК Звезда — {len(users)} чел.",
    )
    return {"sent": True, "count": len(users)}


@router.patch("/bulk", response_model=AdminUsersBulkResult)
async def bulk_update_users(
    payload: AdminUsersBulkUpdate,
    current_user: CurrentUser = Depends(require_role(RoleLevel.ADMIN)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> AdminUsersBulkResult:
    if current_user.user_id is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="User profile is required.")
    user_changes = payload.model_dump(
        include={"role_code", "squad_id", "status_code"},
        exclude_none=True,
    )
    preference_changes = payload.model_dump(
        include={"telegram_enabled", "vk_enabled", "web_push_enabled", "in_app_enabled"},
        exclude_none=True,
    )
    if "role_code" in user_changes:
        if user_changes["role_code"] not in ROLE_LEVELS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown role_code.")
        target_level = ROLE_LEVELS[user_changes["role_code"]]
        if target_level >= current_user.role_level and current_user.role_level < RoleLevel.SUPER_ADMIN:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot assign this role.")
    if "status_code" in user_changes and user_changes["status_code"] not in {"ACTIVE", "ARCHIVED"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown status_code.")
    if "squad_id" in user_changes and await session.get(Squad, user_changes["squad_id"]) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Squad not found.")

    users = list(
        (
            await session.scalars(
                select(User).where(User.id.in_(payload.user_ids)).order_by(User.id).with_for_update()
            )
        ).all()
    )
    if len(users) != len(payload.user_ids):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="One or more users not found.")
    audit_before: list[dict] = []
    audit_after: list[dict] = []
    revoke_user_ids: list[int] = []
    for user in users:
        if user.id == current_user.user_id and any(key in user_changes for key in ("role_code", "status_code")):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot change your own role or status.")
        if ROLE_LEVELS.get(user.role_code, RoleLevel.PUBLIC_USER) >= current_user.role_level and current_user.role_level < RoleLevel.SUPER_ADMIN:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot change an equal or higher role.")
        before = {field: getattr(user, field) for field in user_changes}
        for field, value in user_changes.items():
            setattr(user, field, value)
        if any(field in user_changes for field in ("role_code", "status_code")):
            bump_token_version(user)
            revoke_user_ids.append(user.id)
        user.updated_at = datetime.now(timezone.utc)
        await sync_user_squad_leadership(session, user)

        before_preferences: list[dict] = []
        after_preferences: list[dict] = []
        if preference_changes:
            existing_preferences = {
                item.category_code: item
                for item in (
                    await session.scalars(
                        select(NotificationPreference).where(NotificationPreference.user_id == user.id)
                    )
                ).all()
            }
            for category in NOTIFICATION_CATEGORIES:
                preference = existing_preferences.get(category)
                before_pref = (
                    {field: getattr(preference, field) for field in preference_changes}
                    if preference is not None
                    else None
                )
                if preference is None:
                    preference = NotificationPreference(user_id=user.id, category_code=category)
                    session.add(preference)
                for field, value in preference_changes.items():
                    setattr(preference, field, value)
                preference.updated_at = datetime.now(timezone.utc)
                before_preferences.append({"category_code": category, "values": before_pref})
                after_preferences.append({"category_code": category, "values": preference_changes})
        audit_before.append({"id": user.id, "before": before, "preferences": before_preferences})
        audit_after.append({"id": user.id, "after": user_changes, "preferences": after_preferences})

    audit = await record_audit(
        session,
        user_id=current_user.user_id,
        action_code="admin.users.bulk",
        entity_name="users",
        old_value=audit_before,
        new_value=audit_after,
        comment=f"Bulk update of {len(users)} users",
    )
    await session.flush()
    audit_batch_id = audit.id
    await session.commit()
    for user_id in revoke_user_ids:
        await delete_user_sessions(settings, user_id)
    return AdminUsersBulkResult(affected=len(users), audit_batch_id=audit_batch_id)


@router.patch("/{user_id}", response_model=UserRead)
async def admin_update_user(
    user_id: int,
    payload: UserUpdate,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_PLATOON_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> User:
    user = await session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found.")
    updates = payload.model_dump(exclude_unset=True)
    if "role_code" in updates and updates["role_code"] is not None:
        if updates["role_code"] not in ROLE_LEVELS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown role_code.")
        target_level = ROLE_LEVELS.get(updates["role_code"], RoleLevel.PUBLIC_USER)
        if target_level >= current_user.role_level and current_user.role_level < RoleLevel.SUPER_ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Cannot assign a role equal to or higher than your own.",
            )
    old = model_snapshot(user, list(updates))
    old_status = user.status_code
    old_telegram_id = user.telegram_id
    old_role = user.role_code
    for key, value in updates.items():
        setattr(user, key, value)
    security_changed = user.status_code != old_status or user.telegram_id != old_telegram_id or user.role_code != old_role
    if security_changed:
        bump_token_version(user)
    await sync_user_squad_leadership(session, user)
    user.updated_at = datetime.now(timezone.utc)
    await record_audit(
        session,
        user_id=current_user.user_id,
        action_code="user.admin_update",
        entity_name="users",
        entity_id=user.id,
        old_value=old,
        new_value=updates,
    )
    await session.commit()
    if security_changed:
        await delete_user_sessions(settings, user.id)
    await session.refresh(user)
    return user


@router.patch("/{user_id}/deactivate", response_model=UserRead)
async def deactivate_user(
    user_id: int,
    current_user: CurrentUser = Depends(require_role(RoleLevel.ADMIN)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> User:
    user = await session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found.")
    if user.id == current_user.user_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot deactivate yourself.")
    if current_user.role_level < RoleLevel.SUPER_ADMIN and ROLE_LEVELS.get(user.role_code, 0) >= current_user.role_level:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot deactivate a user with equal or higher role.")
    old_status = user.status_code
    user.status_code = "ARCHIVED"
    user.updated_at = datetime.now(timezone.utc)
    bump_token_version(user)
    await record_audit(
        session,
        user_id=current_user.user_id,
        action_code="user.deactivate",
        entity_name="users",
        entity_id=user.id,
        old_value={"status_code": old_status},
        new_value={"status_code": "ARCHIVED"},
    )
    await session.commit()
    await delete_user_sessions(settings, user.id)
    await session.refresh(user)
    return user
