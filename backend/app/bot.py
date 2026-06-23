from __future__ import annotations

import asyncio
import csv
import io
import logging
import re
from datetime import date, datetime, timedelta, timezone

from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.storage.redis import RedisStorage
from aiogram.types import (
    BotCommand,
    BufferedInputFile,
    CallbackQuery,
    ErrorEvent,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InlineQuery,
    InlineQueryResultArticle,
    InputTextMessageContent,
    KeyboardButton,
    MenuButtonWebApp,
    Message,
    ReplyKeyboardMarkup,
    WebAppInfo,
)
from sqlalchemy import select

from .config import get_settings
from .database import AsyncSessionLocal
from .models import AbsenceReason, Appeal, Attendance, EventResponse, JoinApplication, Normative, NormativeSubmission, Notification, ScheduleEvent, Squad, User
from .roles import RoleLevel, role_level
from .services.auth_security import PasswordPolicyError, bump_token_version, validate_password_policy
from .services.delivery import call_telegram_with_rate_limit
from .services.events import save_event_response as save_event_response_service
from .services.heartbeat import heartbeat_loop
from .services.observability import init_sentry
from .services.normatives import submit_normative as submit_normative_service
from .utils.audit import record_audit
from .utils.channel_link import issue_link_code
from .utils.password import hash_password

logger = logging.getLogger(__name__)
router = Router(name="vpk-zvezda-db-bot")

_CF_URL_RE = re.compile(r'https://[a-z0-9-]+\.trycloudflare\.com')
_CF_LOG_PATH = "/tmp/cf_tunnel.log"
JOIN_STATE_TIMEOUT = timedelta(days=1)
ATTENDANCE_PAGE_SIZE = 8


def _detect_tunnel_url() -> str | None:
    try:
        with open(_CF_LOG_PATH) as f:
            content = f.read()
        matches = _CF_URL_RE.findall(content)
        return matches[-1] if matches else None
    except OSError:
        return None


class AbsenceReasonStates(StatesGroup):
    awaiting_custom = State()


class JoinApplicationStates(StatesGroup):
    full_name = State()
    birth_date = State()
    phone = State()
    motivation = State()
    source = State()
    confirm = State()


class AppealStates(StatesGroup):
    subject = State()
    description = State()
    urgency = State()


class PasswordResetStates(StatesGroup):
    new_password = State()
    confirm_password = State()


ROLE_LABELS = {
    "PUBLIC_USER": "Новый пользователь",
    "CANDIDATE": "Кандидат",
    "USER_PENDING": "Ожидает привязки",
    "PARTICIPANT": "Участник",
    "DEPUTY_SQUAD_COMMANDER": "Заместитель командира отделения",
    "SQUAD_COMMANDER": "Командир отделения",
    "DEPUTY_PLATOON_COMMANDER": "Заместитель командира взвода",
    "PLATOON_COMMANDER": "Командир взвода",
    "ADMIN": "Администратор",
    "SUPER_ADMIN": "Супер-администратор",
}

PARTICIPANT_ROLE_CODES = (
    "PARTICIPANT",
    "DEPUTY_SQUAD_COMMANDER",
    "SQUAD_COMMANDER",
    "DEPUTY_PLATOON_COMMANDER",
    "PLATOON_COMMANDER",
    "ADMIN",
    "SUPER_ADMIN",
)


def parse_birth_date(value: str) -> date | None:
    text = value.strip()
    if text.casefold() in {"-", "нет", "пропустить"}:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("Invalid birth date")


async def find_user(telegram_id: int) -> User | None:
    async with AsyncSessionLocal() as session:
        return await session.scalar(select(User).where(User.telegram_id == telegram_id))


def user_role(user: User | None) -> RoleLevel:
    return role_level(user.role_code if user else "PUBLIC_USER")


def main_keyboard(role: RoleLevel) -> ReplyKeyboardMarkup:
    settings = get_settings()
    rows: list[list[KeyboardButton]] = [[KeyboardButton(text="Меню"), KeyboardButton(text="Расписание"), KeyboardButton(text="Уведомления")]]
    if role >= RoleLevel.DEPUTY_SQUAD_COMMANDER:
        rows.append([KeyboardButton(text="Заявки")])
    if settings.mini_app_url:
        rows.append([KeyboardButton(text="Открыть приложение", web_app=WebAppInfo(url=settings.mini_app_url))])
    if role >= RoleLevel.SUPER_ADMIN:
        rows.append([KeyboardButton(text="Ссылка туннеля")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True, one_time_keyboard=False)


def cancel_keyboard() -> ReplyKeyboardMarkup:
    settings = get_settings()
    rows: list[list[KeyboardButton]] = [[KeyboardButton(text="Отмена")]]
    if settings.mini_app_url:
        rows.append([KeyboardButton(text="Открыть приложение", web_app=WebAppInfo(url=settings.mini_app_url))])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True, one_time_keyboard=False)


def main_menu_inline(role: RoleLevel) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = [
        [
            InlineKeyboardButton(text="Расписание", callback_data="menu:schedule"),
            InlineKeyboardButton(text="Уведомления", callback_data="menu:notifications"),
        ],
    ]
    if role >= RoleLevel.PARTICIPANT:
        rows.append(
            [
                InlineKeyboardButton(text="Нормативы", callback_data="menu:normatives"),
                InlineKeyboardButton(text="Моя явка", callback_data="menu:attendance"),
            ]
        )
        rows.append(
            [
                InlineKeyboardButton(text="Обращение", callback_data="menu:appeal"),
                InlineKeyboardButton(text="Привязать VK", callback_data="menu:vk"),
            ]
        )
    else:
        rows.append([InlineKeyboardButton(text="Заявка на вступление", callback_data="menu:join")])
    settings = get_settings()
    if settings.mini_app_url:
        rows.append([InlineKeyboardButton(text="Открыть приложение", web_app=WebAppInfo(url=settings.mini_app_url))])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def event_keyboard(event_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Приду", callback_data=f"event:{event_id}:COMING"),
                InlineKeyboardButton(text="Не приду", callback_data=f"event:{event_id}:NOT_COMING"),
                InlineKeyboardButton(text="Пока не знаю", callback_data=f"event:{event_id}:MAYBE"),
            ]
        ]
    )


def mini_app_keyboard() -> InlineKeyboardMarkup | None:
    settings = get_settings()
    if not settings.mini_app_url:
        return None
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Открыть приложение", web_app=WebAppInfo(url=settings.mini_app_url))]
        ]
    )


def absence_reasons_keyboard(event_id: int, reasons: list[AbsenceReason]) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text=reason.label, callback_data=f"reason:{event_id}:{reason.id}")]
        for reason in reasons
    ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def save_event_response(
    *,
    event_id: int,
    user_id: int,
    response_code: str,
    absence_reason_id: int | None = None,
    custom_reason: str | None = None,
) -> None:
    async with AsyncSessionLocal() as session:
        await save_event_response_service(
            session,
            event_id=event_id,
            user_id=user_id,
            response_code=response_code,
            absence_reason_id=absence_reason_id,
            custom_reason=custom_reason,
            source_code="BOT",
        )
        await session.commit()


async def ensure_dialog_not_expired(message: Message, state: FSMContext) -> bool:
    data = await state.get_data()
    started_at_raw = data.get("started_at")
    if not started_at_raw:
        return False
    try:
        started_at = datetime.fromisoformat(started_at_raw)
    except (TypeError, ValueError):
        return False
    if started_at.tzinfo is None:
        started_at = started_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) - started_at <= JOIN_STATE_TIMEOUT:
        return False
    await state.clear()
    await message.answer("Диалог устарел и был сброшен. Начните заново нужной командой.", parse_mode=None)
    return True


async def show_main_menu(message: Message, user: User | None = None) -> None:
    user = user or await find_user(message.from_user.id)
    role = user_role(user)
    await message.answer("Выберите раздел:", reply_markup=main_menu_inline(role), parse_mode=None)


@router.message(Command("start", "menu"))
async def start(message: Message, bot: Bot) -> None:
    user = await find_user(message.from_user.id)
    role = user_role(user)

    # Handle deep link parameters
    text = message.text or ""
    if " " in text:
        param = text.split(" ", 1)[1]
        if param.startswith("view"):
            submission_id_str = param[4:]
            try:
                submission_id = int(submission_id_str)
            except (ValueError, TypeError):
                submission_id = None
            if submission_id is not None and role >= RoleLevel.DEPUTY_SQUAD_COMMANDER:
                async with AsyncSessionLocal() as session:
                    submission = await session.get(NormativeSubmission, submission_id)
                if submission is not None:
                    comment = submission.comment or ""
                    if "[TG file_id: " in comment:
                        # Extract TG file_id from comment
                        start_idx = comment.index("[TG file_id: ") + len("[TG file_id: ")
                        end_idx = comment.find("]", start_idx)
                        tg_file_id = comment[start_idx:end_idx] if end_idx != -1 else comment[start_idx:]
                        try:
                            await bot.send_document(message.from_user.id, tg_file_id)
                        except Exception:  # noqa: BLE001
                            logger.exception("Failed to forward TG file for submission_id=%s", submission_id)
                            await message.answer("Не удалось переслать файл.", parse_mode=None)
                    elif submission.file_id is not None:
                        await message.answer("Файл загружен через приложение. Откройте Mini App для просмотра.", parse_mode=None)
                    else:
                        await message.answer("Файл для этой сдачи не найден.", parse_mode=None)
                    return
            elif submission_id is not None:
                await message.answer("Доступ только командирам.", parse_mode=None)
                return

    name = user.full_name if user else message.from_user.full_name
    role_label = ROLE_LABELS.get(user.role_code if user else "PUBLIC_USER", "Пользователь")
    lines = [
        "ВПК «Звезда»",
        "",
        f"Пользователь: {name}",
        f"Должность: {role_label}",
    ]
    if user is None:
        lines += ["", "Если вы в составе — попросите командира привязать ваш Telegram ID."]
        lines += ["Для вступления: /join"]
    await message.answer("\n".join(lines), reply_markup=main_keyboard(role), parse_mode=None)
    await show_main_menu(message, user)


@router.message(F.text.casefold().in_({"меню"}))
async def menu_text(message: Message) -> None:
    await show_main_menu(message)


@router.message(Command("profile"))
async def profile(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user is None:
        await message.answer("Профиль пока не найден. Откройте Mini App и подайте заявку или попросите командира привязать вас.")
        return
    lines = [
        user.full_name,
        f"Должность: {ROLE_LABELS.get(user.role_code, user.role_code)}",
        f"Отделение: {user.squad_id or 'не назначено'}",
        f"Статус: {user.status_code}",
    ]
    await message.answer("\n".join(lines), parse_mode=None)


@router.message(F.text == "Ссылка туннеля")
async def tunnel_url(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.SUPER_ADMIN:
        return
    url = _detect_tunnel_url()
    settings = get_settings()
    lines: list[str] = []
    if url:
        lines.append(f"Текущий URL туннеля:\n{url}")
    else:
        lines.append("URL туннеля не найден в логах cf-tunnel.")
    env_url = settings.mini_app_url
    if env_url and env_url != url:
        lines.append(f"\nВ .env сейчас другой URL:\n{env_url}")
    elif env_url:
        lines.append("\nСовпадает с MINI_APP_URL в .env")
    reply_markup = (
        InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Открыть Mini App", web_app=WebAppInfo(url=url))]
            ]
        )
        if url
        else mini_app_keyboard()
    )
    await message.answer("\n".join(lines), reply_markup=reply_markup, parse_mode=None)


@router.message(Command("cancel"))
@router.message(F.text.casefold().in_({"отмена", "cancel", "стоп"}))
async def cancel_dialog(message: Message, state: FSMContext) -> None:
    current_state = await state.get_state()
    await state.clear()
    user = await find_user(message.from_user.id)
    reply_markup = main_keyboard(user_role(user))
    if current_state:
        await message.answer("Действие отменено.", reply_markup=reply_markup, parse_mode=None)
    else:
        await message.answer("Активного диалога нет. Открыл меню.", reply_markup=reply_markup, parse_mode=None)
        await show_main_menu(message, user)


@router.message(Command("join"))
async def join(message: Message, state: FSMContext) -> None:
    user = await find_user(message.from_user.id)
    role = user_role(user)
    if role >= RoleLevel.PARTICIPANT:
        await message.answer("Вы уже в составе. Откройте личный кабинет через Mini App.", reply_markup=mini_app_keyboard())
        return
    async with AsyncSessionLocal() as session:
        application = await session.scalar(
            select(JoinApplication)
            .where(JoinApplication.telegram_id == message.from_user.id)
            .order_by(JoinApplication.id.desc())
        )
    if application:
        await message.answer(f"Ваша заявка уже есть. Статус: {application.status_code}", reply_markup=mini_app_keyboard())
    else:
        await state.set_state(JoinApplicationStates.full_name)
        await state.update_data(started_at=datetime.now(timezone.utc).isoformat())
        await message.answer(
            "Заполним заявку прямо здесь.\n\nНапишите ФИО кандидата.",
            reply_markup=cancel_keyboard(),
            parse_mode=None,
        )


@router.message(JoinApplicationStates.full_name)
async def join_full_name(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    full_name = (message.text or "").strip()
    if len(full_name) < 2:
        await message.answer("ФИО слишком короткое. Напишите фамилию и имя.")
        return
    await state.update_data(full_name=full_name)
    await state.set_state(JoinApplicationStates.birth_date)
    await message.answer(
        "Дата рождения в формате ДД.ММ.ГГГГ. Если не хотите указывать, напишите «пропустить».",
        reply_markup=cancel_keyboard(),
    )


@router.message(JoinApplicationStates.birth_date)
async def join_birth_date(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    try:
        birth_date = parse_birth_date(message.text or "")
    except ValueError:
        await message.answer("Не понял дату. Пример: 14.02.2010. Можно написать «пропустить».")
        return
    await state.update_data(birth_date=birth_date.isoformat() if birth_date else None)
    await state.set_state(JoinApplicationStates.phone)
    await message.answer("Телефон для связи. Если не хотите указывать, напишите «пропустить».", reply_markup=cancel_keyboard())


@router.message(JoinApplicationStates.phone)
async def join_phone(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    phone = (message.text or "").strip()
    await state.update_data(phone=None if phone.casefold() in {"-", "нет", "пропустить"} else phone)
    await state.set_state(JoinApplicationStates.motivation)
    await message.answer("Почему хотите вступить в ВПК «Звезда»?", reply_markup=cancel_keyboard())


@router.message(JoinApplicationStates.motivation)
async def join_motivation(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    motivation = (message.text or "").strip()
    if len(motivation) < 3:
        await message.answer("Напишите хотя бы коротко, зачем хотите вступить.")
        return
    await state.update_data(motivation_text=motivation)
    await state.set_state(JoinApplicationStates.source)
    await message.answer("Откуда узнали о ВПК? Можно написать «пропустить».", reply_markup=cancel_keyboard())


@router.message(JoinApplicationStates.source)
async def join_source(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    source = (message.text or "").strip()
    await state.update_data(source_text=None if source.casefold() in {"-", "нет", "пропустить"} else source)
    data = await state.get_data()
    lines = [
        "Проверьте заявку:",
        f"ФИО: {data.get('full_name')}",
        f"Дата рождения: {data.get('birth_date') or 'не указана'}",
        f"Телефон: {data.get('phone') or 'не указан'}",
        f"Мотивация: {data.get('motivation_text')}",
        f"Источник: {data.get('source_text') or 'не указан'}",
        "",
        "Отправить заявку? Напишите «да» или «нет».",
    ]
    await state.set_state(JoinApplicationStates.confirm)
    await message.answer("\n".join(lines), reply_markup=cancel_keyboard(), parse_mode=None)


@router.message(JoinApplicationStates.confirm)
async def join_confirm(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    answer = (message.text or "").strip().casefold()
    if answer not in {"да", "нет"}:
        await message.answer("Напишите «да», чтобы отправить, или «нет», чтобы отменить.")
        return
    if answer == "нет":
        await state.clear()
        await message.answer("Заявка отменена. Можно начать заново командой /join.", reply_markup=main_keyboard(RoleLevel.PUBLIC_USER))
        return
    data = await state.get_data()
    async with AsyncSessionLocal() as session:
        existing = await session.scalar(
            select(JoinApplication)
            .where(
                JoinApplication.telegram_id == message.from_user.id,
                JoinApplication.status_code.not_in(["REJECTED", "ARCHIVED", "ACCEPTED"]),
            )
            .order_by(JoinApplication.id.desc())
        )
        if existing:
            await state.clear()
            await message.answer(f"Активная заявка уже есть. Статус: {existing.status_code}", reply_markup=mini_app_keyboard())
            return
        user = await session.scalar(select(User).where(User.telegram_id == message.from_user.id))
        application = JoinApplication(
            telegram_id=message.from_user.id,
            username=message.from_user.username,
            full_name=data["full_name"],
            birth_date=date.fromisoformat(data["birth_date"]) if data.get("birth_date") else None,
            phone=data.get("phone"),
            motivation_text=data.get("motivation_text"),
            source_text=data.get("source_text"),
            consent_given=True,
            status_code="NEW",
        )
        session.add(application)
        if user is None:
            user = User(
                telegram_id=message.from_user.id,
                username=message.from_user.username,
                full_name=data["full_name"],
                birth_date=application.birth_date,
                phone=application.phone,
                role_code="CANDIDATE",
                status_code="ACTIVE",
                linked_at=datetime.now(timezone.utc),
            )
            session.add(user)
        elif user.role_code == "PUBLIC_USER":
            user.full_name = data["full_name"]
            user.birth_date = application.birth_date
            user.phone = application.phone
            user.role_code = "CANDIDATE"
            user.updated_at = datetime.now(timezone.utc)
        await session.flush()
        await record_audit(
            session,
            user_id=user.id,
            action_code="join.application.create_bot",
            entity_name="join_applications",
            entity_id=application.id,
            new_value={"telegram_id": message.from_user.id, "source": "bot"},
        )
        commanders = list((await session.scalars(
            select(User).where(
                User.status_code == "ACTIVE",
                User.role_code.in_(("PLATOON_COMMANDER", "DEPUTY_PLATOON_COMMANDER")),
            )
        )).all())
        for commander in commanders:
            session.add(Notification(
                user_id=commander.id,
                type_code="NEW_APPLICATION",
                title="Новая заявка",
                body=f"{data['full_name']} подал(а) заявку на вступление через бот.",
                entity_name="join_applications",
                entity_id=application.id,
                send_to_tg=True,
            ))
        await session.commit()
    await state.clear()
    await message.answer(
        "Заявка отправлена. Командиры уведомлены.\nСтатус — в приложении",
        reply_markup=main_keyboard(RoleLevel.CANDIDATE),
    )


@router.message(Command("help"))
async def help_command(message: Message) -> None:
    await message.answer(
        "Помощь по боту ВПК «Звезда»:\n\n"
        "/start или /menu — главное меню.\n"
        "/schedule — ближайшие занятия. Под уведомлением о занятии нажмите «Приду», «Не приду» или «Пока не знаю».\n"
        "/normatives — активные нормативы. Чтобы сдать норматив, просто пришлите фото, видео или документ в бот и выберите норматив.\n"
        "/attendance — ваши отметки посещаемости.\n"
        "/notifications — последние уведомления.\n"
        "/appeal — обращение командиру без Mini App.\n"
        "/vk — код для привязки ВКонтакте.\n"
        "/resetpassword — код сброса и смена пароля через Telegram.\n"
        "/join — заявка на вступление.\n"
        "/profile — ваш профиль.\n"
        "/cancel — отменить текущий диалог.\n\n"
        "Командирам доступны заявки, экспорт состава, проверка нормативов и отметка явки через кнопки в уведомлениях.",
        parse_mode=None,
    )


@router.message(Command("vk"))
async def vk_link(message: Message) -> None:
    await send_vk_link_code(message, message.from_user.id)


async def send_vk_link_code(message: Message, telegram_id: int) -> None:
    user = await find_user(telegram_id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Привязка VK доступна подтверждённым участникам состава.", parse_mode=None)
        return
    async with AsyncSessionLocal() as session:
        db_user = await session.get(User, user.id)
        if db_user is None:
            await message.answer("Профиль не найден. Обратитесь к командиру.", parse_mode=None)
            return
        code, expires_at = await issue_link_code(session, db_user.id, channel="VK")
        await record_audit(
            session,
            user_id=db_user.id,
            action_code="vk.link_code.issue_bot",
            entity_name="users",
            entity_id=db_user.id,
        )
        await session.commit()
    settings = get_settings()
    lines = [
        "Код привязки VK:",
        code,
        "",
        f"Действует до {expires_at.strftime('%d.%m %H:%M')} UTC.",
        "Отправьте этот код в VK-бот одним сообщением.",
    ]
    if settings.vk_bot_url:
        lines.append(f"\nVK-бот: {settings.vk_bot_url}")
    await message.answer("\n".join(lines), parse_mode=None)


@router.message(Command("resetpassword"))
async def reset_password_start(message: Message, state: FSMContext) -> None:
    user = await find_user(message.from_user.id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Сброс пароля доступен подтверждённым участникам состава.", parse_mode=None)
        return
    async with AsyncSessionLocal() as session:
        db_user = await session.get(User, user.id)
        if db_user is None:
            await message.answer("Профиль не найден. Обратитесь к командиру.", parse_mode=None)
            return
        code, expires_at = await issue_link_code(session, db_user.id, channel="PASSWORD_RESET")
        await record_audit(
            session,
            user_id=db_user.id,
            action_code="auth.password.reset_code.issue_bot",
            entity_name="users",
            entity_id=db_user.id,
        )
        await session.commit()
    await state.set_state(PasswordResetStates.new_password)
    await state.update_data(started_at=datetime.now(timezone.utc).isoformat(), user_id=user.id)
    await message.answer(
        "Код сброса пароля для входа на сайте:\n"
        f"{code}\n\n"
        f"Действует до {expires_at.strftime('%d.%m %H:%M')} UTC.\n"
        "Можно ввести код на странице входа или задать новый пароль прямо здесь: напишите новый пароль.",
        reply_markup=cancel_keyboard(),
        parse_mode=None,
    )


@router.message(PasswordResetStates.new_password)
async def reset_password_new_password(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    password = (message.text or "").strip()
    user = await find_user(message.from_user.id)
    if user is None:
        await state.clear()
        await message.answer("Профиль не найден. Обратитесь к командиру.", parse_mode=None)
        return
    try:
        validate_password_policy(password, telegram_id=user.telegram_id)
    except PasswordPolicyError as exc:
        await message.answer(str(exc), parse_mode=None)
        return
    await state.update_data(new_password=password)
    await state.set_state(PasswordResetStates.confirm_password)
    await message.answer("Повторите новый пароль для подтверждения.", reply_markup=cancel_keyboard(), parse_mode=None)


@router.message(PasswordResetStates.confirm_password)
async def reset_password_confirm(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    data = await state.get_data()
    password = str(data.get("new_password") or "")
    if (message.text or "").strip() != password:
        await state.set_state(PasswordResetStates.new_password)
        await state.update_data(new_password=None)
        await message.answer("Пароли не совпали. Напишите новый пароль ещё раз.", reply_markup=cancel_keyboard())
        return
    user_id = int(data.get("user_id") or 0)
    async with AsyncSessionLocal() as session:
        user = await session.get(User, user_id)
        if user is None:
            await state.clear()
            await message.answer("Профиль не найден. Обратитесь к командиру.", parse_mode=None)
            return
        user.password_hash = hash_password(password)
        user.password_set_at = datetime.now(timezone.utc)
        user.updated_at = datetime.now(timezone.utc)
        user.failed_login_count = 0
        user.locked_until = None
        bump_token_version(user)
        await record_audit(
            session,
            user_id=user.id,
            action_code="auth.password.reset_bot",
            entity_name="users",
            entity_id=user.id,
        )
        await session.commit()
    await state.clear()
    await message.answer("Пароль обновлён. Старые сессии сайта отозваны.", reply_markup=main_keyboard(user_role(user)), parse_mode=None)


@router.callback_query(F.data.startswith("menu:"))
async def menu_callback(callback: CallbackQuery, state: FSMContext) -> None:
    action = (callback.data or "").split(":", 1)[1]
    message = callback.message
    if message is None:
        await callback.answer()
        return
    user = await find_user(callback.from_user.id)
    role = user_role(user)
    if action == "schedule":
        if user is None or role < RoleLevel.PARTICIPANT:
            await message.answer("Расписание доступно после подтверждения участия.")
        else:
            await _send_schedule_with_batch(message, user)
    elif action == "notifications":
        if user is None or role < RoleLevel.PARTICIPANT:
            await message.answer("Уведомления доступны после подтверждения участия.")
        else:
            await send_notifications_text(message, user)
    elif action == "normatives":
        if user is None or role < RoleLevel.PARTICIPANT:
            await message.answer("Нормативы доступны после подтверждения участия.")
        else:
            await send_normatives_text(message, user)
    elif action == "attendance":
        if user is None or role < RoleLevel.PARTICIPANT:
            await message.answer("Посещаемость доступна после подтверждения участия.")
        else:
            await send_attendance_text(message, user)
    elif action == "appeal":
        await start_appeal_dialog(message, state, callback.from_user.id)
    elif action == "vk":
        await send_vk_link_code(message, callback.from_user.id)
    elif action == "join":
        await message.answer("Чтобы подать заявку, отправьте команду /join.", parse_mode=None)
    else:
        await callback.answer("Раздел пока недоступен.", show_alert=True)
        return
    await callback.answer()


@router.message(Command("admin"))
async def admin(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.ADMIN:
        await message.answer("Команда доступна только администраторам.")
        return
    await message.answer(
        "Резервные команды:\n"
        "/broadcast текст — рассылка всем активным пользователям\n"
        "/schedule — проверить ближайшие события\n"
        "Основное управление доступно в Mini App.",
        reply_markup=mini_app_keyboard(),
        parse_mode=None,
    )


@router.message(Command("broadcast"))
async def broadcast(message: Message, bot: Bot) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.ADMIN:
        await message.answer("Команда доступна только администраторам.")
        return
    text = (message.text or "").partition(" ")[2].strip()
    if not text:
        await message.answer("Использование: /broadcast текст сообщения")
        return
    async with AsyncSessionLocal() as session:
        users = list(
            (
                await session.scalars(
                    select(User).where(User.telegram_id.is_not(None), User.status_code == "ACTIVE")
                )
            ).all()
        )
    sent = 0
    for recipient in users:
        try:
            await call_telegram_with_rate_limit(lambda recipient=recipient: bot.send_message(recipient.telegram_id, text))
            sent += 1
        except Exception:  # noqa: BLE001
            logger.exception("Failed to send broadcast to user_id=%s", recipient.id)
    await message.answer(f"Рассылка отправлена: {sent}/{len(users)}")


@router.message(Command("schedule"))
@router.message(F.text.casefold().in_({"расписание"}))
async def schedule(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Расписание доступно после подтверждения участия.")
        return
    await _send_schedule_with_batch(message, user)


@router.callback_query(F.data.startswith("event:"))
async def event_response(callback: CallbackQuery, state: FSMContext) -> None:
    parts = (callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer("Некорректный ответ.", show_alert=True)
        return
    event_id = int(parts[1])
    response_code = parts[2]
    user = await find_user(callback.from_user.id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await callback.answer("Нужна привязка к составу.", show_alert=True)
        return
    async with AsyncSessionLocal() as session:
        event = await session.get(ScheduleEvent, event_id)
        if event is None:
            await callback.answer("Событие не найдено.", show_alert=True)
            return
        if event.response_deadline_at and datetime.now(timezone.utc) > event.response_deadline_at:
            await callback.answer("Дедлайн ответа уже прошёл.", show_alert=True)
            return
        if response_code == "NOT_COMING" and event.requires_response:
            reasons = list(
                (
                    await session.scalars(
                        select(AbsenceReason).where(AbsenceReason.is_active.is_(True)).order_by(AbsenceReason.sort_order)
                    )
                ).all()
            )
            await callback.message.edit_text(
                "Выберите причину отсутствия:",
                reply_markup=absence_reasons_keyboard(event_id, reasons),
            )
            await callback.answer()
            return
    maybe_deadline: datetime | None = None
    if response_code == "MAYBE":
        async with AsyncSessionLocal() as session2:
            ev = await session2.get(ScheduleEvent, event_id)
            if ev and ev.response_deadline_at:
                maybe_deadline = ev.response_deadline_at
                reminder_at = maybe_deadline - timedelta(hours=1)
                if reminder_at > datetime.now(timezone.utc):
                    session2.add(
                        Notification(
                            user_id=user.id,
                            type_code="SCHEDULE",
                            title="Напоминание об ответе",
                            body=f"Вы ещё не дали окончательный ответ на «{ev.title}». Дедлайн: {maybe_deadline.strftime('%d.%m %H:%M')} UTC.",
                            entity_name="schedule_events",
                            entity_id=event_id,
                            send_to_tg=True,
                        )
                    )
                    await session2.commit()
    await save_event_response(event_id=event_id, user_id=user.id, response_code=response_code)
    label = {"COMING": "буду", "MAYBE": "уточню позже"}.get(response_code, "ответ сохранён")
    extra = ""
    if response_code == "MAYBE" and maybe_deadline:
        extra = f" Напомним до {maybe_deadline.strftime('%d.%m %H:%M')}."
    if callback.message:
        await callback.message.edit_text(f"Ответ записан: {label}.{extra}")
    await callback.answer("Ответ сохранён.")


@router.callback_query(F.data.startswith("reason:"))
async def absence_reason(callback: CallbackQuery, state: FSMContext) -> None:
    parts = (callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer("Некорректная причина.", show_alert=True)
        return
    event_id = int(parts[1])
    reason_id = int(parts[2])
    user = await find_user(callback.from_user.id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await callback.answer("Нужна привязка к составу.", show_alert=True)
        return
    async with AsyncSessionLocal() as session:
        reason = await session.get(AbsenceReason, reason_id)
        if reason is None:
            await callback.answer("Причина не найдена.", show_alert=True)
            return
        if reason.requires_comment:
            await state.set_state(AbsenceReasonStates.awaiting_custom)
            await state.update_data(event_id=event_id, reason_id=reason_id, user_id=user.id, reason_label=reason.label)
            await callback.message.edit_text("Напишите причину одним сообщением.")
            await callback.answer()
            return
    await save_event_response(event_id=event_id, user_id=user.id, response_code="NOT_COMING", absence_reason_id=reason_id)
    if callback.message:
        await callback.message.edit_text(f"Ответ записан: не приду. Причина: {reason.label}")
    await callback.answer("Причина сохранена.")


@router.message(AbsenceReasonStates.awaiting_custom)
async def absence_custom_reason(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    text = (message.text or "").strip()
    if not text:
        await message.answer("Напишите причину текстом.")
        return
    await save_event_response(
        event_id=int(data["event_id"]),
        user_id=int(data["user_id"]),
        response_code="NOT_COMING",
        absence_reason_id=int(data["reason_id"]),
        custom_reason=text,
    )
    await state.clear()
    await message.answer(f"Ответ записан: не приду. Причина: {data.get('reason_label')}: {text}", parse_mode=None)


@router.message(Command("attendance"))
async def attendance(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Посещаемость доступна после подтверждения участия.")
        return
    await send_attendance_text(message, user)


async def send_attendance_text(message: Message, user: User) -> None:
    STATUS_LABELS_ATT = {
        "PRESENT": "Присутствовал",
        "ABSENT": "Отсутствовал",
        "LATE": "Опоздал",
        "EXCUSED": "Уважительная",
        "SICK": "Больничный",
        "RELEASED": "Освобождён",
    }
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Attendance, ScheduleEvent.title)
            .join(ScheduleEvent, ScheduleEvent.id == Attendance.event_id, isouter=True)
            .where(Attendance.user_id == user.id)
            .order_by(Attendance.updated_at.desc())
            .limit(10)
        )
        rows = result.all()
    if not rows:
        await message.answer("Отметок посещаемости пока нет.")
        return
    lines = ["Моя посещаемость:"]
    for att, title in rows:
        date_str = att.updated_at.strftime("%d.%m") if att.updated_at else ""
        status_label = STATUS_LABELS_ATT.get(att.status_code, att.status_code)
        lines.append(f"• {title or 'Занятие'} {date_str}: {status_label}")
    await message.answer("\n".join(lines), parse_mode=None)


@router.message(Command("normatives"))
async def normatives(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Нормативы доступны после подтверждения участия.")
        return
    await send_normatives_text(message, user)


async def send_normatives_text(message: Message, user: User) -> None:
    async with AsyncSessionLocal() as session:
        rows = list(
            (
                await session.scalars(
                    select(Normative)
                    .where(Normative.is_active.is_(True))
                    .where((Normative.squad_id.is_(None)) | (Normative.squad_id == user.squad_id))
                    .order_by(Normative.deadline_at.nullslast(), Normative.created_at.desc())
                    .limit(10)
                )
            ).all()
        )
    if not rows:
        await message.answer("Активных нормативов пока нет.")
        return
    lines = ["Нормативы:"]
    for item in rows:
        deadline = item.deadline_at.strftime("%d.%m %H:%M") if item.deadline_at else "без дедлайна"
        lines.append(f"• {item.title}: {deadline}")
    await message.answer("\n".join(lines), parse_mode=None)


@router.message(Command("notifications"))
@router.message(F.text.casefold().in_({"уведомления"}))
async def notifications(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Уведомления доступны после подтверждения участия.")
        return
    await send_notifications_text(message, user)


async def send_notifications_text(message: Message, user: User) -> None:
    async with AsyncSessionLocal() as session:
        rows = list(
            (
                await session.scalars(
                    select(Notification)
                    .where(Notification.user_id == user.id)
                    .order_by(Notification.is_pinned.desc(), Notification.created_at.desc())
                    .limit(10)
                )
            ).all()
        )
    if not rows:
        await message.answer("Уведомлений пока нет.")
        return
    unread = [r for r in rows if not r.is_read]
    lines = [f"Уведомления ({len(unread)} новых):"]
    for item in rows:
        prefix = "новое:" if not item.is_read else "•"
        lines.append(f"{prefix} {item.title}")
    await message.answer("\n".join(lines), parse_mode=None)


@router.message(Command("appeal"))
async def appeal_start(message: Message, state: FSMContext) -> None:
    await start_appeal_dialog(message, state, message.from_user.id)


async def start_appeal_dialog(message: Message, state: FSMContext, telegram_id: int) -> None:
    user = await find_user(telegram_id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await message.answer("Обращения доступны после подтверждения участия.", parse_mode=None)
        return
    await state.set_state(AppealStates.subject)
    await state.update_data(started_at=datetime.now(timezone.utc).isoformat(), user_id=user.id)
    await message.answer("Напишите тему обращения.", reply_markup=cancel_keyboard(), parse_mode=None)


@router.message(AppealStates.subject)
async def appeal_subject(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    subject = (message.text or "").strip()
    if len(subject) < 3:
        await message.answer("Тема слишком короткая. Напишите чуть подробнее.")
        return
    await state.update_data(subject=subject[:255])
    await state.set_state(AppealStates.description)
    await message.answer("Опишите ситуацию одним сообщением.", reply_markup=cancel_keyboard(), parse_mode=None)


@router.message(AppealStates.description)
async def appeal_description(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    description = (message.text or "").strip()
    if len(description) < 5:
        await message.answer("Описание слишком короткое. Добавьте деталей.")
        return
    await state.update_data(description=description)
    await state.set_state(AppealStates.urgency)
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Обычная", callback_data="appeal_urgency:NORMAL"),
                InlineKeyboardButton(text="Срочная", callback_data="appeal_urgency:HIGH"),
            ],
            [InlineKeyboardButton(text="Очень срочно", callback_data="appeal_urgency:URGENT")],
        ]
    )
    await message.answer("Выберите срочность:", reply_markup=keyboard, parse_mode=None)


@router.message(AppealStates.urgency)
async def appeal_urgency_text(message: Message, state: FSMContext) -> None:
    if await ensure_dialog_not_expired(message, state):
        return
    low = (message.text or "").casefold()
    urgency = "URGENT" if "очень" in low else "HIGH" if "сроч" in low else "NORMAL"
    await create_appeal_from_state(message, state, urgency)


@router.callback_query(F.data.startswith("appeal_urgency:"))
async def appeal_urgency_callback(callback: CallbackQuery, state: FSMContext) -> None:
    urgency = (callback.data or "").split(":", 1)[1]
    if urgency not in {"LOW", "NORMAL", "HIGH", "URGENT"}:
        urgency = "NORMAL"
    if callback.message:
        await create_appeal_from_state(callback.message, state, urgency)
    await callback.answer("Обращение отправлено.")


async def create_appeal_from_state(message: Message, state: FSMContext, urgency: str) -> None:
    data = await state.get_data()
    user_id = int(data.get("user_id") or 0)
    async with AsyncSessionLocal() as session:
        user = await session.get(User, user_id)
        if user is None:
            await state.clear()
            await message.answer("Профиль не найден. Обратитесь к командиру.", parse_mode=None)
            return
        appeal = Appeal(
            author_user_id=user.id,
            is_anonymous=False,
            subject=str(data.get("subject") or "Обращение")[:255],
            category_code="OTHER",
            description=str(data.get("description") or ""),
            urgency_code=urgency,
            status_code="CREATED",
        )
        session.add(appeal)
        await session.flush()
        await record_audit(
            session,
            user_id=user.id,
            action_code="appeal.create_bot",
            entity_name="appeals",
            entity_id=appeal.id,
            new_value={"source": "telegram", "urgency_code": urgency},
        )
        commanders = list(
            (
                await session.scalars(
                    select(User).where(
                        User.status_code == "ACTIVE",
                        User.role_code.in_(("DEPUTY_PLATOON_COMMANDER", "PLATOON_COMMANDER", "ADMIN", "SUPER_ADMIN")),
                    )
                )
            ).all()
        )
        for commander in commanders:
            session.add(
                Notification(
                    user_id=commander.id,
                    type_code="APPEAL",
                    title=f"Новое обращение: {appeal.subject}",
                    body=f"{user.full_name} отправил обращение через Telegram. Срочность: {urgency}.",
                    entity_name="appeals",
                    entity_id=appeal.id,
                    send_to_tg=True,
                )
            )
        await session.commit()
    await state.clear()
    await message.answer("Обращение отправлено. Ответ придёт в уведомления.", reply_markup=main_keyboard(user_role(user)), parse_mode=None)


@router.message(F.text.casefold().in_({"заявки"}))
async def cmd_applications(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user_role(user) < RoleLevel.DEPUTY_SQUAD_COMMANDER:
        await message.answer("Доступ только командирам.")
        return
    async with AsyncSessionLocal() as session:
        from .models import JoinApplication
        apps = list((await session.scalars(
            select(JoinApplication)
            .where(JoinApplication.status_code.not_in(["ACCEPTED", "REJECTED", "ARCHIVED"]))
            .order_by(JoinApplication.created_at.desc())
            .limit(10)
        )).all())
    if not apps:
        await message.answer("Активных заявок нет.", reply_markup=mini_app_keyboard())
        return
    STATUS_LABELS = {
        "NEW": "Новая", "INVITED_NORMATIVES": "На нормативах",
        "REVIEWING": "На рассмотрении", "NEEDS_INFO": "Нужна информация",
    }
    lines = [f"Заявки ({len(apps)}):"]
    for app in apps:
        status = STATUS_LABELS.get(app.status_code, app.status_code)
        lines.append(f"• {app.full_name} — {status}")
    lines.append("\nДля работы с заявками откройте приложение:")
    await message.answer("\n".join(lines), reply_markup=mini_app_keyboard(), parse_mode=None)


@router.callback_query(F.data.startswith("norm_review:"))
async def normative_review_callback(callback: CallbackQuery) -> None:
    parts = (callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    try:
        submission_id, status_code = int(parts[1]), parts[2]
    except ValueError:
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    user = await find_user(callback.from_user.id)
    if user is None or user_role(user) < RoleLevel.DEPUTY_SQUAD_COMMANDER:
        await callback.answer("Только для командиров.", show_alert=True)
        return
    STATUS_LABELS = {"ACCEPTED": "Принято", "REJECTED": "Отклонено", "NEEDS_REDO": "На доработку"}
    if status_code not in STATUS_LABELS:
        await callback.answer("Некорректный статус.", show_alert=True)
        return
    status_label = STATUS_LABELS.get(status_code, status_code)
    async with AsyncSessionLocal() as session:
        submission = await session.get(NormativeSubmission, submission_id)
        if submission is None:
            await callback.answer("Сдача не найдена.", show_alert=True)
            return
        submitter = await session.get(User, submission.user_id)
        normative = await session.get(Normative, submission.normative_id)
        norm_title = normative.title if normative else f"норматив #{submission.normative_id}"
        submission.status_code = status_code
        submission.reviewed_by_id = user.id
        from datetime import datetime, timezone
        submission.reviewed_at = datetime.now(timezone.utc)
        submission.updated_at = datetime.now(timezone.utc)
        submitter_name = submitter.full_name if submitter else "участник"
        if submitter:
            body_parts = [f"Норматив: «{norm_title}»"]
            if status_code == "ACCEPTED":
                body_parts.append("Ваша сдача принята!")
            elif status_code == "REJECTED":
                body_parts.append("Ваша сдача отклонена.")
            else:
                body_parts.append("Требуется пересдача.")
            session.add(Notification(
                user_id=submitter.id,
                type_code="NORMATIVE",
                title=f"{status_label}: {norm_title}",
                body="\n".join(body_parts),
                entity_name="normative_submissions",
                entity_id=submission.id,
                send_to_tg=True,
            ))
        await session.commit()
    if callback.message:
        try:
            await callback.message.edit_text(
                f"{status_label} — {norm_title}\nУчастник: {submitter_name}",
                parse_mode=None,
            )
        except Exception:  # noqa: BLE001
            logger.exception("Failed to edit normative review message for submission_id=%s", submission_id)
    await callback.answer(f"Статус: {status_label}")


# ──────────────────────── /schedule batch reply ─────────────────────────────


def batch_events_keyboard(events: list[ScheduleEvent]) -> InlineKeyboardMarkup:
    """Show 'Answer all' shortcut when there are 2+ upcoming unanswered events."""
    rows = [
        [
            InlineKeyboardButton(text="На все", callback_data="batch:COMING"),
            InlineKeyboardButton(text="Ни на одно", callback_data="batch:NOT_COMING"),
        ],
        [InlineKeyboardButton(text="По одному", callback_data="batch:ONE_BY_ONE")],
    ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _upcoming_unanswered_events(session, user: User, limit: int = 7) -> list[ScheduleEvent]:
    now = datetime.now(timezone.utc)
    events = list(
        (
            await session.scalars(
                select(ScheduleEvent)
                .where(
                    ScheduleEvent.start_datetime >= now,
                    ScheduleEvent.status_code != "CANCELLED",
                    (ScheduleEvent.squad_id.is_(None)) | (ScheduleEvent.squad_id == user.squad_id),
                    ScheduleEvent.requires_response.is_(True),
                )
                .order_by(ScheduleEvent.start_datetime)
                .limit(limit)
            )
        ).all()
    )
    if not events:
        return []
    response_rows = (
        await session.execute(
            select(EventResponse.event_id).where(
                EventResponse.user_id == user.id,
                EventResponse.event_id.in_([event.id for event in events]),
            )
        )
    ).all()
    answered_ids = {event_id for (event_id,) in response_rows}
    return [event for event in events if event.id not in answered_ids]


@router.callback_query(F.data.startswith("batch:"))
async def batch_response(callback: CallbackQuery) -> None:
    action = (callback.data or "").split(":")[1]
    user = await find_user(callback.from_user.id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await callback.answer("Нужна привязка к составу.", show_alert=True)
        return
    now = datetime.now(timezone.utc)
    async with AsyncSessionLocal() as session:
        events = await _upcoming_unanswered_events(session, user)
        if action == "ONE_BY_ONE":
            await callback.answer()
            if callback.message:
                await callback.message.delete()
            for event in events:
                start_at = event.start_datetime.strftime("%d.%m %H:%M")
                place = f"\nМесто: {event.place}" if event.place else ""
                await callback.message.answer(
                    f"{event.title}\n{start_at}{place}",
                    reply_markup=event_keyboard(event.id),
                    parse_mode=None,
                )
            return
        if action not in {"COMING", "NOT_COMING"}:
            await callback.answer("Некорректный ответ.", show_alert=True)
            return
        if not events:
            if callback.message:
                await callback.message.edit_text("Нет занятий без ответа.")
            await callback.answer("Нечего обновлять.")
            return
        saved = 0
        for event in events:
            existing = await session.scalar(
                select(EventResponse).where(EventResponse.event_id == event.id, EventResponse.user_id == user.id)
            )
            if existing is None:
                existing = EventResponse(event_id=event.id, user_id=user.id)
                session.add(existing)
            existing.response_code = action
            existing.responded_at = now
            existing.source_code = "BOT"
            saved += 1
        await session.commit()
    label = "буду на всех" if action == "COMING" else "не приду ни на одно"
    if callback.message:
        await callback.message.edit_text(f"Ответ записан: {label}. Затронуто {saved} событий.")
    await callback.answer("Готово!")


# ──────────────────────── commander attendance via bot ─────────────────────


def attendance_mark_keyboard(
    *,
    event_id: int,
    users: list[User],
    existing_status: dict[int, str],
    page: int,
    total: int,
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for user in users:
        name = user.full_name.split()[0] if user.full_name else f"#{user.id}"
        status_code = existing_status.get(user.id)
        present_label = f"{name}: был" if status_code != "PRESENT" else f"{name}: был ✓"
        late_label = "опоздал" if status_code != "LATE" else "опоздал ✓"
        absent_label = "нет" if status_code != "ABSENT" else "нет ✓"
        rows.append(
            [
                InlineKeyboardButton(text=present_label[:32], callback_data=f"attmark:{event_id}:{user.id}:PRESENT:{page}"),
                InlineKeyboardButton(text=late_label, callback_data=f"attmark:{event_id}:{user.id}:LATE:{page}"),
                InlineKeyboardButton(text=absent_label, callback_data=f"attmark:{event_id}:{user.id}:ABSENT:{page}"),
            ]
        )
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="Назад", callback_data=f"attpage:{event_id}:{page - 1}"))
    if (page + 1) * ATTENDANCE_PAGE_SIZE < total:
        nav.append(InlineKeyboardButton(text="Дальше", callback_data=f"attpage:{event_id}:{page + 1}"))
    if nav:
        rows.append(nav)
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _attendance_event_context(session, commander: User, event_id: int) -> tuple[ScheduleEvent | None, str | None]:
    event = await session.get(ScheduleEvent, event_id)
    if event is None:
        return None, "Событие не найдено."
    role = user_role(commander)
    if role < RoleLevel.DEPUTY_SQUAD_COMMANDER:
        return None, "Отмечать явку могут только командиры."
    if role < RoleLevel.DEPUTY_PLATOON_COMMANDER and commander.squad_id is None:
        return None, "У вас не назначено отделение."
    if role < RoleLevel.DEPUTY_PLATOON_COMMANDER and event.squad_id is not None and event.squad_id != commander.squad_id:
        return None, "Это событие другого отделения."
    return event, None


async def send_attendance_mark_page(message: Message, commander: User, event_id: int, page: int = 0) -> None:
    async with AsyncSessionLocal() as session:
        event, error = await _attendance_event_context(session, commander, event_id)
        if error or event is None:
            await message.answer(error or "Событие недоступно.")
            return
        statement = select(User).where(User.status_code == "ACTIVE", User.role_code.in_(PARTICIPANT_ROLE_CODES))
        if event.squad_id is not None:
            statement = statement.where(User.squad_id == event.squad_id)
        elif user_role(commander) < RoleLevel.DEPUTY_PLATOON_COMMANDER and commander.squad_id is not None:
            statement = statement.where(User.squad_id == commander.squad_id)
        participants = list((await session.scalars(statement.order_by(User.full_name))).all())
        total = len(participants)
        page = max(0, page)
        page_users = participants[page * ATTENDANCE_PAGE_SIZE:(page + 1) * ATTENDANCE_PAGE_SIZE]
        if not page_users:
            await message.answer("Некого отмечать по этому событию.")
            return
        rows = (
            await session.execute(
                select(Attendance.user_id, Attendance.status_code).where(
                    Attendance.event_id == event.id,
                    Attendance.user_id.in_([user.id for user in page_users]),
                )
            )
        ).all()
    status_by_user = {user_id: status_code for user_id, status_code in rows}
    text = (
        f"Явка: {event.title}\n"
        f"Страница {page + 1}/{max(1, (total + ATTENDANCE_PAGE_SIZE - 1) // ATTENDANCE_PAGE_SIZE)}"
    )
    await message.answer(
        text,
        reply_markup=attendance_mark_keyboard(
            event_id=event_id,
            users=page_users,
            existing_status=status_by_user,
            page=page,
            total=total,
        ),
        parse_mode=None,
    )


@router.callback_query(F.data.startswith("attendance:"))
async def attendance_open_callback(callback: CallbackQuery) -> None:
    parts = (callback.data or "").split(":")
    try:
        event_id = int(parts[1])
        page = int(parts[2]) if len(parts) > 2 else 0
    except (IndexError, ValueError):
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    commander = await find_user(callback.from_user.id)
    if commander is None:
        await callback.answer("Профиль не найден.", show_alert=True)
        return
    if callback.message:
        await send_attendance_mark_page(callback.message, commander, event_id, page)
    await callback.answer()


@router.callback_query(F.data.startswith("attpage:"))
async def attendance_page_callback(callback: CallbackQuery) -> None:
    parts = (callback.data or "").split(":")
    try:
        event_id = int(parts[1])
        page = int(parts[2])
    except (IndexError, ValueError):
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    commander = await find_user(callback.from_user.id)
    if commander is None:
        await callback.answer("Профиль не найден.", show_alert=True)
        return
    if callback.message:
        await callback.message.delete()
        await send_attendance_mark_page(callback.message, commander, event_id, page)
    await callback.answer()


@router.callback_query(F.data.startswith("attmark:"))
async def attendance_mark_callback(callback: CallbackQuery) -> None:
    parts = (callback.data or "").split(":")
    try:
        event_id = int(parts[1])
        participant_id = int(parts[2])
        status_code = parts[3]
        page = int(parts[4])
    except (IndexError, ValueError):
        await callback.answer("Некорректная отметка.", show_alert=True)
        return
    if status_code not in {"PRESENT", "LATE", "ABSENT"}:
        await callback.answer("Некорректный статус.", show_alert=True)
        return
    commander = await find_user(callback.from_user.id)
    if commander is None:
        await callback.answer("Профиль не найден.", show_alert=True)
        return
    now = datetime.now(timezone.utc)
    async with AsyncSessionLocal() as session:
        db_commander = await session.get(User, commander.id)
        if db_commander is None:
            await callback.answer("Профиль не найден.", show_alert=True)
            return
        event, error = await _attendance_event_context(session, db_commander, event_id)
        participant = await session.get(User, participant_id)
        if error or event is None or participant is None:
            await callback.answer(error or "Участник не найден.", show_alert=True)
            return
        if event.squad_id is not None and participant.squad_id != event.squad_id:
            await callback.answer("Участник из другого отделения.", show_alert=True)
            return
        attendance_row = await session.scalar(
            select(Attendance).where(Attendance.event_id == event_id, Attendance.user_id == participant_id)
        )
        old_status = attendance_row.status_code if attendance_row else None
        if attendance_row is None:
            attendance_row = Attendance(event_id=event_id, user_id=participant_id)
            session.add(attendance_row)
        attendance_row.status_code = status_code
        attendance_row.marked_by_user_id = db_commander.id
        attendance_row.marked_at = now
        attendance_row.updated_at = now
        await session.flush()
        await record_audit(
            session,
            user_id=db_commander.id,
            action_code="attendance.mark_bot",
            entity_name="attendance",
            entity_id=attendance_row.id,
            old_value={"status_code": old_status},
            new_value={"event_id": event_id, "user_id": participant_id, "status_code": status_code},
        )
        await session.commit()
    if callback.message:
        await callback.message.delete()
        await send_attendance_mark_page(callback.message, commander, event_id, page)
    await callback.answer("Явка обновлена.")


# ──────────────────────── /schedule — enriched with batch ───────────────────


async def _send_schedule_with_batch(message: Message, user: User) -> None:
    now = datetime.now(timezone.utc)
    async with AsyncSessionLocal() as session:
        events = list(
            (
                await session.scalars(
                    select(ScheduleEvent)
                    .where(
                        ScheduleEvent.start_datetime >= now,
                        ScheduleEvent.status_code != "CANCELLED",
                        (ScheduleEvent.squad_id.is_(None)) | (ScheduleEvent.squad_id == user.squad_id),
                    )
                    .order_by(ScheduleEvent.start_datetime)
                    .limit(7)
                )
            ).all()
        )
        if not events:
            await message.answer("Ближайших событий пока нет.")
            return
        unanswered = []
        for event in events:
            resp = await session.scalar(
                select(EventResponse).where(EventResponse.event_id == event.id, EventResponse.user_id == user.id)
            )
            if resp is None and event.requires_response:
                unanswered.append(event)
        # If 2+ unanswered events — offer batch
        if len(unanswered) >= 2:
            titles = "\n".join(
                f"• {e.title} {e.start_datetime.strftime('%d.%m %H:%M')}" for e in unanswered[:5]
            )
            await message.answer(
                f"Вы не ответили на {len(unanswered)} занятий:\n{titles}\n\nОтветить сразу на все?",
                reply_markup=batch_events_keyboard(unanswered),
                parse_mode=None,
            )
        else:
            for event in events:
                start_at = event.start_datetime.strftime("%d.%m %H:%M")
                place = f"\nМесто: {event.place}" if event.place else ""
                resp = await session.scalar(
                    select(EventResponse).where(EventResponse.event_id == event.id, EventResponse.user_id == user.id)
                )
                resp_label = ""
                if resp:
                    resp_label = {"COMING": " Вы идёте", "NOT_COMING": " Вы не идёте", "MAYBE": " Вы уточняете"}.get(resp.response_code, "")
                await message.answer(
                    f"{event.title}\n{start_at}{place}{resp_label}",
                    reply_markup=event_keyboard(event.id) if event.requires_response else None,
                    parse_mode=None,
                )


# ──────────────────────── video/photo/doc → normative ───────────────────────


class NormativeFileStates(StatesGroup):
    awaiting_normative_choice = State()


def normative_choice_keyboard(normatives: list[Normative]) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text=n.title[:40], callback_data=f"norm_submit:{n.id}")]
        for n in normatives[:8]
    ]
    rows.append([InlineKeyboardButton(text="Это не для норматива", callback_data="norm_submit:cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.message(F.video | F.photo | F.document)
async def handle_file_message(message: Message, state: FSMContext) -> None:
    user = await find_user(message.from_user.id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        return  # silently ignore from non-participants
    async with AsyncSessionLocal() as session:
        normatives = list(
            (
                await session.scalars(
                    select(Normative)
                    .where(
                        Normative.is_active.is_(True),
                        (Normative.squad_id.is_(None)) | (Normative.squad_id == user.squad_id),
                    )
                    .order_by(Normative.deadline_at.nullslast())
                    .limit(8)
                )
            ).all()
        )
    if not normatives:
        return  # no active normatives, don't bother

    # Store file info in FSM state
    file_id: str | None = None
    if message.video:
        file_id = message.video.file_id
    elif message.photo:
        file_id = message.photo[-1].file_id  # largest size
    elif message.document:
        file_id = message.document.file_id

    if not file_id:
        return

    await state.set_state(NormativeFileStates.awaiting_normative_choice)
    await state.update_data(file_id=file_id, user_id=user.id)
    await message.answer(
        "Это файл для сдачи норматива?\nВыберите норматив или отмените:",
        reply_markup=normative_choice_keyboard(normatives),
        parse_mode=None,
    )


@router.callback_query(F.data.startswith("norm_submit:"))
async def normative_file_submit(callback: CallbackQuery, state: FSMContext) -> None:
    norm_id_str = (callback.data or "").split(":")[1]
    if norm_id_str == "cancel":
        await state.clear()
        await callback.message.edit_text("Понял, файл не для норматива.")
        await callback.answer()
        return
    norm_id = int(norm_id_str)
    data = await state.get_data()
    file_id: str = data.get("file_id", "")
    stored_user_id: int = data.get("user_id", 0)

    user = await find_user(callback.from_user.id)
    if user is None or user.id != stored_user_id:
        await callback.answer("Ошибка сессии.", show_alert=True)
        await state.clear()
        return

    async with AsyncSessionLocal() as session:
        normative = await session.get(Normative, norm_id)
        if not normative:
            await callback.answer("Норматив не найден.", show_alert=True)
            return
        submission = await submit_normative_service(
            session,
            normative=normative,
            submitter=user,
            status_code="PENDING",
            comment=f"[TG file_id: {file_id}]",
            file_ids=None,
            audit_action_code="normative_submission.submit_via_bot",
            audit_value={"normative_id": norm_id, "telegram_file_id": file_id, "source": "bot"},
            notification_body=f"{user.full_name} прислал файл на проверку по нормативу «{normative.title}».",
            notification_scope="submitter_squad",
        )
        await session.commit()
        await session.refresh(submission)
        normative_title = normative.title
    await state.clear()
    if callback.message:
        await callback.message.edit_text(
            f"Файл принят на проверку по нормативу «{normative_title}».\nКомандир получил уведомление.",
            parse_mode=None,
        )
    await callback.answer("Сдача отправлена!")


# ──────────────────────── export ────────────────────────────────────────────

EXPORT_COLUMNS = [
    "ID", "Telegram ID", "Имя", "Роль", "Отделение",
    "Статус", "Телефон", "Дата рождения", "Дата привязки",
]


async def _build_export_rows(user: User) -> tuple[list[list[str]], dict[int, str]]:
    async with AsyncSessionLocal() as session:
        stmt = select(User).where(User.role_code != "PUBLIC_USER").order_by(User.full_name)
        if role_level(user.role_code) < RoleLevel.DEPUTY_PLATOON_COMMANDER and user.squad_id:
            stmt = stmt.where(User.squad_id == user.squad_id)
        users = list((await session.scalars(stmt)).all())
        squads = {s.id: s.name for s in (await session.scalars(select(Squad))).all()}
    rows = []
    for u in users:
        rows.append([
            str(u.id or ""),
            str(u.telegram_id or ""),
            u.full_name or "",
            u.role_code or "",
            squads.get(u.squad_id, "—") if u.squad_id else "—",
            u.status_code or "",
            u.phone or "",
            u.birth_date.strftime("%d.%m.%Y") if u.birth_date else "",
            u.linked_at.strftime("%d.%m.%Y") if getattr(u, "linked_at", None) else "",
        ])
    return rows, squads


@router.message(Command("export"))
async def cmd_export(message: Message) -> None:
    user = await find_user(message.from_user.id)
    if user is None or user_role(user) < RoleLevel.DEPUTY_SQUAD_COMMANDER:
        await message.answer("Нет прав для экспорта данных.", parse_mode=None)
        return
    await message.answer(
        "Выберите формат экспорта состава:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Excel (.xlsx)", callback_data="export:xlsx"),
            InlineKeyboardButton(text="CSV", callback_data="export:csv"),
        ]]),
        parse_mode=None,
    )


@router.callback_query(F.data.startswith("export:"))
async def cb_export(callback: CallbackQuery, bot: Bot) -> None:
    fmt = (callback.data or "").split(":")[1]
    user = await find_user(callback.from_user.id)
    if user is None or user_role(user) < RoleLevel.DEPUTY_SQUAD_COMMANDER:
        await callback.answer("Нет прав.", show_alert=True)
        return

    if callback.message:
        await callback.message.edit_text("Формирую файл, подождите...")
    await callback.answer()

    rows, _ = await _build_export_rows(user)

    if fmt == "csv":
        buf = io.StringIO()
        buf.write("﻿")  # BOM for Excel
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(EXPORT_COLUMNS)
        for row in rows:
            writer.writerow(row)
        data = buf.getvalue().encode("utf-8")
        await bot.send_document(
            callback.from_user.id,
            BufferedInputFile(data, filename="vpk-roster.csv"),
            caption=f"Состав ВПК Звезда ({len(rows)} чел.)",
        )
    elif fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            await bot.send_message(callback.from_user.id, "Сервер не поддерживает XLSX. Используйте CSV.", parse_mode=None)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Состав"
        header_fill = PatternFill(fill_type="solid", fgColor="1a2f5a")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(EXPORT_COLUMNS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append(row)
        buf_b = io.BytesIO()
        wb.save(buf_b)
        buf_b.seek(0)
        await bot.send_document(
            callback.from_user.id,
            BufferedInputFile(buf_b.read(), filename="vpk-roster.xlsx"),
            caption=f"Состав ВПК Звезда ({len(rows)} чел.)",
        )

    if callback.message:
        await callback.message.edit_text(f"Файл отправлен ({len(rows)} участников).")


# ──────────────────────── inline mode ───────────────────────────────────────


@router.inline_query()
async def inline_schedule(query: InlineQuery) -> None:
    """@vpkbot [текст] — показать расписание в любом чате."""
    user = await find_user(query.from_user.id)
    if user is None or user_role(user) < RoleLevel.PARTICIPANT:
        await query.answer(
            [InlineQueryResultArticle(
                id="no_access",
                title="Нет доступа",
                input_message_content=InputTextMessageContent(message_text="Сначала зарегистрируйтесь в ВПК «Звезда»."),
            )],
            cache_time=10,
        )
        return

    now = datetime.now(timezone.utc)
    async with AsyncSessionLocal() as session:
        events = list(
            (
                await session.scalars(
                    select(ScheduleEvent)
                    .where(
                        ScheduleEvent.start_datetime >= now,
                        (ScheduleEvent.squad_id.is_(None)) | (ScheduleEvent.squad_id == user.squad_id),
                        ScheduleEvent.status_code != "CANCELLED",
                    )
                    .order_by(ScheduleEvent.start_datetime)
                    .limit(5)
                )
            ).all()
        )

    results = []
    for event in events:
        start_str = event.start_datetime.strftime("%d.%m %H:%M")
        place_str = f" · {event.place}" if event.place else ""
        text = f"{event.title}\n{start_str}{place_str}"
        results.append(
            InlineQueryResultArticle(
                id=str(event.id),
                title=f"{event.title}",
                description=f"{start_str}{place_str}",
                input_message_content=InputTextMessageContent(message_text=text),
            )
        )
    if not results:
        results.append(
            InlineQueryResultArticle(
                id="empty",
                title="Событий нет",
                input_message_content=InputTextMessageContent(message_text="Ближайших событий не запланировано."),
            )
        )
    await query.answer(results, cache_time=60)


@router.errors()
async def bot_error_handler(event: ErrorEvent) -> bool:
    logger.exception("Unhandled Telegram bot error", exc_info=event.exception)
    update = event.update
    message = getattr(update, "message", None) or getattr(update, "callback_query", None)
    try:
        if getattr(message, "message", None):
            await message.message.answer("Произошла ошибка. Попробуйте позже.", parse_mode=None)
        elif hasattr(message, "answer"):
            await message.answer("Произошла ошибка. Попробуйте позже.", parse_mode=None)
    except Exception:  # noqa: BLE001
        logger.exception("Failed to notify user about bot error")
    return True


def build_storage(settings):
    if settings.redis_url:
        return RedisStorage.from_url(
            settings.redis_url,
            state_ttl=int(JOIN_STATE_TIMEOUT.total_seconds()),
            data_ttl=int(JOIN_STATE_TIMEOUT.total_seconds()),
        )
    return MemoryStorage()


async def main() -> None:
    logging.basicConfig(level=logging.INFO)
    settings = get_settings()
    init_sentry(settings, service_name="telegram_bot")
    asyncio.create_task(heartbeat_loop(settings.bot_heartbeat_path))
    if settings.dryrun:
        logger.info("Starting VPK Zvezda Telegram bot in DRYRUN mode; polling is disabled")
        await asyncio.Event().wait()
    bot = Bot(settings.bot_token)
    if settings.mini_app_url:
        await bot.set_chat_menu_button(
            menu_button=MenuButtonWebApp(
                text="Открыть приложение",
                web_app=WebAppInfo(url=settings.mini_app_url),
            )
        )
    await bot.set_my_commands([
        BotCommand(command="start", description="Главное меню"),
        BotCommand(command="schedule", description="Расписание занятий"),
        BotCommand(command="notifications", description="Мои уведомления"),
        BotCommand(command="normatives", description="Нормативы"),
        BotCommand(command="attendance", description="Моя посещаемость"),
        BotCommand(command="appeal", description="Обращение командиру"),
        BotCommand(command="vk", description="Привязать VK"),
        BotCommand(command="resetpassword", description="Сбросить пароль сайта"),
        BotCommand(command="cancel", description="Отменить диалог"),
        BotCommand(command="join", description="Заявка на вступление"),
        BotCommand(command="profile", description="Мой профиль"),
        BotCommand(command="export", description="Выгрузка состава (CSV/Excel)"),
        BotCommand(command="help", description="Помощь"),
    ])
    dispatcher = Dispatcher(storage=build_storage(settings))
    dispatcher.include_router(router)
    logger.info("Starting VPK Zvezda Telegram bot")
    await dispatcher.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
