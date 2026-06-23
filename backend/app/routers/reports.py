from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import Settings, get_settings
from ..database import get_db_session
from ..dependencies.auth import CurrentUser, require_role
from ..models import Appeal, Attendance, AttendanceGrade, EventResponse, JoinApplication, NormativeSubmission, ScheduleEvent, Squad, User
from ..roles import CONFIRMED_ROLES, RoleLevel
from ..schemas.core import ReportSummary

router = APIRouter(prefix="/reports", tags=["reports"])

ATTENDANCE_EXPORT_COLUMNS = [
    "Дата занятия",
    "Событие",
    "Место",
    "ФИО",
    "Telegram",
    "Отделение",
    "Статус",
    "Отмечено",
]

ATTENDANCE_STATUS_LABELS = {
    "PRESENT": "Присутствовал",
    "ABSENT": "Отсутствовал",
    "LATE": "Опоздал",
    "EXCUSED": "Уважительная",
    "SICK": "Больничный",
    "RELEASED": "Освобождён",
    "NOT_MARKED": "Не отмечен",
}

# Short labels shown inside the attendance matrix cells (табель).
MATRIX_STATUS_SHORT = {
    "PRESENT": "Был",
    "ABSENT": "Не был",
    "LATE": "Опоздал",
    "EXCUSED": "Уваж.",
    "SICK": "Болел",
    "RELEASED": "Освоб.",
}
# When several events fall on one day, the status with the highest priority wins.
MATRIX_STATUS_PRIORITY = {
    "PRESENT": 6,
    "LATE": 5,
    "EXCUSED": 4,
    "SICK": 3,
    "RELEASED": 2,
    "ABSENT": 1,
    "NOT_MARKED": 0,
}
MATRIX_DROPDOWN = "Был,Не был,Опоздал,Уваж.,Болел,Освоб."


def _format_dt(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M")


async def _attendance_export_rows(
    current_user: CurrentUser,
    session: AsyncSession,
    squad_id: int | None = None,
) -> list[list[str]]:
    statement = (
        select(
            ScheduleEvent.start_datetime,
            ScheduleEvent.title,
            ScheduleEvent.place,
            User.full_name,
            User.username,
            User.squad_id,
            Attendance.status_code,
            Attendance.marked_at,
        )
        .join(ScheduleEvent, ScheduleEvent.id == Attendance.event_id)
        .join(User, User.id == Attendance.user_id)
        .order_by(ScheduleEvent.start_datetime.desc(), User.full_name)
    )
    if squad_id is not None:
        statement = statement.where(User.squad_id == squad_id)
    elif current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER:
        statement = statement.where(User.squad_id == current_user.squad_id)

    squad_map = {squad.id: squad.name for squad in (await session.scalars(select(Squad))).all()}
    rows = (await session.execute(statement)).all()
    return [
        [
            _format_dt(start_datetime),
            title,
            place or "",
            full_name,
            f"@{username}" if username else "",
            squad_map.get(user_squad_id, ""),
            ATTENDANCE_STATUS_LABELS.get(status_code, status_code),
            _format_dt(marked_at),
        ]
        for start_datetime, title, place, full_name, username, user_squad_id, status_code, marked_at in rows
    ]


@router.get("/attendance", response_model=ReportSummary)
async def attendance_report(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> ReportSummary:
    statement = select(Attendance.status_code, func.count(Attendance.id)).join(User, User.id == Attendance.user_id)
    if squad_id is not None:
        statement = statement.where(User.squad_id == squad_id)
    elif current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER:
        statement = statement.where(User.squad_id == current_user.squad_id)
    rows = (await session.execute(statement.group_by(Attendance.status_code))).all()
    return ReportSummary(title="Посещаемость", items=[{"status_code": key, "count": count} for key, count in rows])


@router.get("/grades", response_model=ReportSummary)
async def grades_report(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> ReportSummary:
    statement = (
        select(AttendanceGrade.grade_value, func.count(AttendanceGrade.id))
        .join(Attendance, Attendance.id == AttendanceGrade.attendance_id)
        .join(User, User.id == Attendance.user_id)
    )
    if squad_id is not None:
        statement = statement.where(User.squad_id == squad_id)
    elif current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER:
        statement = statement.where(User.squad_id == current_user.squad_id)
    rows = (await session.execute(statement.group_by(AttendanceGrade.grade_value))).all()
    return ReportSummary(title="Оценки", items=[{"grade_value": key, "count": count} for key, count in rows])


@router.get("/normatives", response_model=ReportSummary)
async def normatives_report(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> ReportSummary:
    statement = (
        select(NormativeSubmission.status_code, func.count(NormativeSubmission.id))
        .join(User, User.id == NormativeSubmission.user_id)
    )
    if squad_id is not None:
        statement = statement.where(User.squad_id == squad_id)
    elif current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER:
        statement = statement.where(User.squad_id == current_user.squad_id)
    rows = (await session.execute(statement.group_by(NormativeSubmission.status_code))).all()
    return ReportSummary(title="Нормативы", items=[{"status_code": key, "count": count} for key, count in rows])


@router.get("/applications", response_model=ReportSummary)
async def applications_report(
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_PLATOON_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> ReportSummary:
    rows = (
        await session.execute(
            select(JoinApplication.status_code, func.count(JoinApplication.id)).group_by(JoinApplication.status_code)
        )
    ).all()
    return ReportSummary(title="Заявки", items=[{"status_code": key, "count": count} for key, count in rows])


@router.get("/export")
async def export_report(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_PLATOON_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    attendance = await attendance_report(squad_id, current_user, session)
    grades = await grades_report(squad_id, current_user, session)
    normatives = await normatives_report(squad_id, current_user, session)
    applications = await applications_report(current_user, session)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["section", "key", "value", "count"])
    for section in [attendance, grades, normatives, applications]:
        for item in section.items:
            key, value = next(((k, v) for k, v in item.items() if k != "count"), ("item", ""))
            writer.writerow([section.title, key, value, item.get("count", "")])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="vpk-report.csv"'},
    )


@router.post("/export/send", status_code=200)
async def export_report_send(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_PLATOON_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Generate report CSV and send it to the requester via Telegram bot DM."""
    from ..background import _get_bot
    from aiogram.types import BufferedInputFile

    attendance = await attendance_report(squad_id, current_user, session)
    grades = await grades_report(squad_id, current_user, session)
    normatives = await normatives_report(squad_id, current_user, session)
    applications = await applications_report(current_user, session)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["section", "key", "value", "count"])
    for section in [attendance, grades, normatives, applications]:
        for item in section.items:
            key, value = next(((k, v) for k, v in item.items() if k != "count"), ("item", ""))
            writer.writerow([section.title, key, value, item.get("count", "")])
    data = output.getvalue().encode("utf-8")
    bot = _get_bot(settings)
    await bot.send_document(
        current_user.telegram_id,
        BufferedInputFile(data, filename="vpk-report.csv"),
        caption="Сводный отчёт ВПК Звезда",
    )
    return {"sent": True}


@router.post("/attendance/export.csv/send", status_code=200)
async def export_attendance_csv_send(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_PLATOON_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Generate detailed attendance CSV and send it to the requester via Telegram bot DM."""
    from ..background import _get_bot
    from aiogram.types import BufferedInputFile

    rows = await _attendance_export_rows(current_user, session, squad_id)
    output = io.StringIO()
    output.write("﻿")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(ATTENDANCE_EXPORT_COLUMNS)
    writer.writerows(rows)
    bot = _get_bot(settings)
    await bot.send_document(
        current_user.telegram_id,
        BufferedInputFile(output.getvalue().encode("utf-8"), filename="attendance.csv"),
        caption=f"Посещаемость ВПК Звезда — {len(rows)} отметок",
    )
    return {"sent": True, "count": len(rows)}


@router.post("/attendance/export.xlsx/send", status_code=200)
async def export_attendance_xlsx_send(
    squad_id: int | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_PLATOON_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Generate detailed attendance XLSX and send it to the requester via Telegram bot DM."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="openpyxl not installed.") from exc
    from ..background import _get_bot
    from aiogram.types import BufferedInputFile

    rows = await _attendance_export_rows(current_user, session, squad_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"
    header_fill = PatternFill(fill_type="solid", fgColor="1A2F5A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(ATTENDANCE_EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 44)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    bot = _get_bot(settings)
    await bot.send_document(
        current_user.telegram_id,
        BufferedInputFile(buf.read(), filename="attendance.xlsx"),
        caption=f"Посещаемость ВПК Звезда — {len(rows)} отметок",
    )
    return {"sent": True, "count": len(rows)}


@router.get("/activity-feed", response_model=list[dict])
async def activity_feed(
    limit: int = 40,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> list[dict]:
    """Recent activity feed for commanders: responses, submissions, appeals, attendance marks."""
    now = datetime.now(timezone.utc)
    since = now - timedelta(days=7)
    squad_filter = current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER and current_user.squad_id is not None

    feed: list[dict] = []

    # Event responses
    resp_rows = (
        await session.execute(
            select(EventResponse.response_code, EventResponse.responded_at, User.full_name, ScheduleEvent.title)
            .join(User, User.id == EventResponse.user_id)
            .join(ScheduleEvent, ScheduleEvent.id == EventResponse.event_id)
            .where(
                EventResponse.responded_at >= since,
                (User.squad_id == current_user.squad_id) if squad_filter else True,
            )
            .order_by(EventResponse.responded_at.desc())
            .limit(20)
        )
    ).all()
    for rc, ts, name, event_title in resp_rows:
        label = {"COMING": "ответил «Приду»", "NOT_COMING": "ответил «Не приду»", "MAYBE": "ответил «Пока не знаю»"}.get(rc, "ответил")
        feed.append({"type": "response", "text": f"{name} {label} на «{event_title}»", "created_at": ts.isoformat()})

    # Normative submissions
    subm_rows = (
        await session.execute(
            select(NormativeSubmission.status_code, NormativeSubmission.submitted_at, User.full_name)
            .join(User, User.id == NormativeSubmission.user_id)
            .where(
                NormativeSubmission.submitted_at >= since,
                (User.squad_id == current_user.squad_id) if squad_filter else True,
            )
            .order_by(NormativeSubmission.submitted_at.desc())
            .limit(15)
        )
    ).all()
    for sc, ts, name in subm_rows:
        feed.append({"type": "normative", "text": f"{name} сдал норматив (статус: {sc})", "created_at": ts.isoformat()})

    # Appeals created
    appeal_rows = (
        await session.execute(
            select(Appeal.subject, Appeal.urgency_code, Appeal.created_at, User.full_name)
            .join(User, User.id == Appeal.author_user_id)
            .where(
                Appeal.created_at >= since,
                (User.squad_id == current_user.squad_id) if squad_filter else True,
            )
            .order_by(Appeal.created_at.desc())
            .limit(10)
        )
    ).all()
    for subject, urgency, ts, name in appeal_rows:
        urgency_label = {"URGENT": "срочное", "HIGH": "важное", "NORMAL": "обычное", "LOW": "низкий приоритет"}.get(urgency, "обычное")
        feed.append({"type": "appeal", "text": f"{name} отправил обращение ({urgency_label}): «{subject}»", "created_at": ts.isoformat()})

    # Recent attendance marks
    att_rows = (
        await session.execute(
            select(Attendance.status_code, Attendance.marked_at, User.full_name, ScheduleEvent.title)
            .join(User, User.id == Attendance.user_id)
            .join(ScheduleEvent, ScheduleEvent.id == Attendance.event_id)
            .where(
                Attendance.marked_at >= since,
                Attendance.marked_at.is_not(None),
                (User.squad_id == current_user.squad_id) if squad_filter else True,
            )
            .order_by(Attendance.marked_at.desc())
            .limit(15)
        )
    ).all()
    status_label = {"PRESENT": "присутствовал", "ABSENT": "отсутствовал", "LATE": "опоздал", "EXCUSED": "уважительная причина", "SICK": "болел"}
    for sc, ts, name, event_title in att_rows:
        label = status_label.get(sc, sc)
        feed.append({"type": "attendance", "text": f"{name} — {label} на «{event_title}»", "created_at": ts.isoformat()})

    # Sort by created_at desc and limit
    feed.sort(key=lambda x: x["created_at"], reverse=True)
    return feed[:limit]


@router.get("/attendance/matrix.xlsx")
async def export_attendance_matrix(
    squad_id: int | None = None,
    month: str | None = None,
    current_user: CurrentUser = Depends(require_role(RoleLevel.DEPUTY_SQUAD_COMMANDER)),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    """Attendance timesheet: rows = members grouped by squad, columns = days that had events.

    Filled cells come from the system; empty ones carry a dropdown so commanders can
    fill them by hand in Excel (Был / Не был / Опоздал / Уваж. / Болел / Освоб.).
    """
    try:
        import openpyxl
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="openpyxl not installed.") from exc

    # Period = a calendar month (default: current).
    today = date.today()
    if month:
        try:
            year_str, month_str = month.split("-", 1)
            period_start = date(int(year_str), int(month_str), 1)
        except (ValueError, TypeError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="month must be YYYY-MM.") from exc
    else:
        period_start = today.replace(day=1)
    period_end = (period_start.replace(day=28) + timedelta(days=4)).replace(day=1)  # first day of next month
    start_dt = datetime(period_start.year, period_start.month, period_start.day, tzinfo=timezone.utc)
    end_dt = datetime(period_end.year, period_end.month, period_end.day, tzinfo=timezone.utc)

    # Squad scope: below platoon-deputy a commander only sees their own squad.
    restricted = current_user.role_level < RoleLevel.DEPUTY_PLATOON_COMMANDER
    scope_squad_id = squad_id if squad_id is not None else (current_user.squad_id if restricted else None)

    # Events in the period (only days with events become columns).
    ev_stmt = select(ScheduleEvent.id, ScheduleEvent.start_datetime).where(
        ScheduleEvent.start_datetime >= start_dt,
        ScheduleEvent.start_datetime < end_dt,
        ScheduleEvent.status_code != "CANCELLED",
    )
    if scope_squad_id is not None:
        ev_stmt = ev_stmt.where((ScheduleEvent.squad_id.is_(None)) | (ScheduleEvent.squad_id == scope_squad_id))
    ev_rows = (await session.execute(ev_stmt)).all()
    event_date = {eid: dt.date() for eid, dt in ev_rows}
    dates = sorted(set(event_date.values()))

    # Aggregate one status per (user, day).
    status_map: dict[tuple[int, date], str] = {}
    if event_date:
        att_rows = (
            await session.execute(
                select(Attendance.user_id, Attendance.event_id, Attendance.status_code).where(
                    Attendance.event_id.in_(list(event_date.keys()))
                )
            )
        ).all()
        for uid, eid, sc in att_rows:
            day = event_date.get(eid)
            if day is None:
                continue
            key = (uid, day)
            prev = status_map.get(key)
            if prev is None or MATRIX_STATUS_PRIORITY.get(sc, 0) > MATRIX_STATUS_PRIORITY.get(prev, 0):
                status_map[key] = sc

    # Members grouped by squad (squads first, "no squad" last), then by name.
    u_stmt = (
        select(User)
        .where(User.status_code == "ACTIVE", User.role_code.in_(CONFIRMED_ROLES))
        .order_by(User.squad_id.nullslast(), User.full_name)
    )
    if scope_squad_id is not None:
        u_stmt = u_stmt.where(User.squad_id == scope_squad_id)
    users = list((await session.scalars(u_stmt)).all())

    squads = {s.id: s for s in (await session.scalars(select(Squad))).all()}
    squad_leads: dict[int, str] = {}
    for squad in squads.values():
        if squad.commander_user_id:
            squad_leads[squad.commander_user_id] = "командир"
        if squad.deputy_user_id:
            squad_leads.setdefault(squad.deputy_user_id, "зам")

    groups: dict[int | None, list[User]] = {}
    for user in users:
        groups.setdefault(user.squad_id, []).append(user)

    # ── build workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"
    name_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="1A2F5A")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12, color="1A2F5A")
    center = Alignment(horizontal="center", vertical="center")
    note_font = Font(italic=True, color="808080", size=9)

    first_col = 3  # A=ФИО, B=примечание, C…=даты
    date_labels = [d.strftime("%d.%m.%Y") for d in dates]
    last_col = first_col + len(dates) - 1

    dv = None
    if dates:
        dv = DataValidation(type="list", formula1=f'"{MATRIX_DROPDOWN}"', allow_blank=True)
        ws.add_data_validation(dv)

    row = 1
    member_ranges: list[str] = []
    for squad_key, members in groups.items():
        squad_name = squads[squad_key].name if squad_key in squads else "Без отделения"
        title_cell = ws.cell(row=row, column=1, value=f"{squad_name} - {len(members)} чел")
        title_cell.font = title_font
        for i, label in enumerate(date_labels):
            c = ws.cell(row=row, column=first_col + i, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        row += 1
        block_start = row
        for member in members:
            name_cell = ws.cell(row=row, column=1, value=member.full_name)
            name_cell.font = name_font
            lead = squad_leads.get(member.id)
            if lead:
                note = ws.cell(row=row, column=2, value=lead)
                note.font = note_font
            for i, day in enumerate(dates):
                label = MATRIX_STATUS_SHORT.get(status_map.get((member.id, day), ""), "")
                cell = ws.cell(row=row, column=first_col + i, value=label or None)
                cell.alignment = center
            row += 1
        if members and dates:
            rng = f"{get_column_letter(first_col)}{block_start}:{get_column_letter(last_col)}{row - 1}"
            dv.add(rng)
            member_ranges.append(rng)
        row += 1  # blank row between squads

    # Conditional formatting: colour the status words (also works for hand-picked values).
    cf_rules = [
        ("Был", "C6EFCE", "006100"),
        ("Не был", "FFC7CE", "9C0006"),
        ("Опоздал", "FFEB9C", "9C6500"),
        ("Уваж.", "DDEBF7", "1F4E78"),
        ("Болел", "DDEBF7", "1F4E78"),
        ("Освоб.", "DDEBF7", "1F4E78"),
    ]
    for rng in member_ranges:
        for word, bg, fg in cf_rules:
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="equal", formula=[f'"{word}"'], fill=PatternFill(fill_type="solid", fgColor=bg), font=Font(color=fg, bold=True)),
            )

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 10
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(first_col + i)].width = 12
    ws.freeze_panes = "C1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attendance-{period_start.strftime('%Y-%m')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
